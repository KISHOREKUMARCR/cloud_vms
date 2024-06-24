from django.shortcuts import render,redirect,get_object_or_404
from .models import *
from .forms import *
from django.views import View 
from django.views.generic.edit import UpdateView
from accounts.models import *
from django.conf import settings
from django.db.models import Q
from django.db import IntegrityError
import psycopg2
from psycopg2 import sql
from urllib.parse import quote
from django.utils import timezone
import sqlite3
import openpyxl
import psycopg2
from django.shortcuts import render
import json
import requests
from django.http import FileResponse
# import paramiko
import datetime
from django.core.mail import send_mail
from accounts.views import *

from .filters import *
from accounts.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
##########
import shutil
from django.conf import settings
from django.templatetags import static
####
# from .filters import*
from django.forms import inlineformset_factory
from django.http import JsonResponse,HttpResponse, Http404,HttpResponseRedirect

### gdrive
from google.oauth2.credentials import Credentials
from google.auth import impersonated_credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from google.oauth2 import service_account
from oauth2client.service_account import ServiceAccountCredentials
from google.auth.transport.requests import Request

###
from googleapiclient.http import MediaFileUpload
from django.shortcuts import render
# from google.oauth2.credentials import Credentials
# from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
import os


import google.oauth2.credentials
import google_auth_oauthlib.flow
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

## resumable uploads to gdrive
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
from googleapiclient.errors import ResumableUploadError
from io import BytesIO
from django.contrib import messages
from datetime import datetime
import pandas as pd

from django.http import (HttpResponse, HttpResponseBadRequest, 
                         HttpResponseForbidden)

from django.db.models.functions import TruncDate
import io

from openpyxl import Workbook
from accounts.models import UserAccount

# def upload_cloud_uri(request):
#     if request.method == 'POST':
#         try:
#             file = request.FILES['file']

#             # Check the file extension
#             if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
#                 df = pd.read_excel(file, dtype=str)
#             elif file.name.endswith('.csv'):
#                 df = pd.read_csv(file, dtype=str)
#             else:
#                 messages.error(request, "Unsupported file type")
#                 return render(request, 'templates/dms/upload_excel_data.html')

#             required_columns = [ 'company_name', 'project_name', 'location_name', 'video_start_time', 'video_end_time', 'onedrive_url', 'userid', 'camera_angle']

#             # Check if all required columns are in the DataFrame
#             if not all(column.strip()  in df.columns for column in required_columns):
#                 missing_columns = [column for column in required_columns if column not in df.columns]
#                 messages.error(request, f"Missing required columns: {', '.join(missing_columns)}")
#                 return render(request, 'templates/dms/upload_excel_data.html')

#             # Iterate over rows in the DataFrame and create CloudURI instances
#             for index, row in df.iterrows():
#                 CloudURI.objects.create(
#                     userid=row.get('userid'),
#                     company_name=row.get('company_name'),
#                     project_name=row.get('project_name'),
#                     location_name=row.get('location_name'),
#                     video_start_time=row.get('video_start_time'),
#                     video_end_time=row.get('video_end_time'),
#                     camera_angle=row.get('camera_angle'),
#                     onedrive_url=row.get('onedrive_url'),
#                 )

#             messages.success(request, "Data has been successfully uploaded to the CloudURI model")

#         except Exception as e:
#             messages.error(request, f"Error processing file: {str(e)}")

#     return render(request, 'templates/dms/upload_excel_data.html')

# def upload_cloud_uri(request):
#     if request.method == 'POST':
#         try:
#             file = request.FILES['file']
#             if not file:
#                 messages.error(request, "No file uploaded")
#                 return render(request, 'templates/dms/upload_excel_data.html')

#             # Check the file extension
#             if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
#                 df = pd.read_excel(file, dtype=str)
#             elif file.name.endswith('.csv'):
#                 df = pd.read_csv(file, dtype=str)
#             else:
#                 messages.error(request, "Unsupported file type")
#                 return render(request, 'templates/dms/upload_excel_data.html')

#             required_columns = [ 'company_name', 'project_name', 'location_name', 'video_start_time', 'video_end_time', 'onedrive_url', 'userid', 'camera_angle']

#             # Check if all required columns are in the DataFrame
#             if not all(column.strip()  in df.columns for column in required_columns):
#                 missing_columns = [column for column in required_columns if column not in df.columns]
#                 messages.error(request, f"Missing required columns: {', '.join(missing_columns)}")
#                 return render(request, 'templates/dms/upload_excel_data.html')

#             # Iterate over rows in the DataFrame and create CloudURI instances
#             for index, row in df.iterrows():            
#                 NewCloudURI.objects.create(
#                     userid=row.get('userid'),
#                     company_name=row.get('company_name'),
#                     project_name=row.get('project_name'),
#                     location_name=row.get('location_name'),
#                     video_start_time=row.get('video_start_time'),
#                     video_end_time=row.get('video_end_time'),
#                     camera_angle=row.get('camera_angle'),
#                     onedrive_url=row.get('onedrive_url'),
#                     )
#             messages.success(request, "Data has been successfully uploaded to the CloudURI model")

#         except Exception as e:
#             messages.error(request, f"Error processing file: {str(e)}")

#     return render(request, 'templates/dms/upload_excel_data.html')



def upload_cloud_uri(request):
    if request.method == 'POST':
        try:
            file = request.FILES.get('file')
            if not file:
                messages.error(request, "No file uploaded")
                return render(request, 'templates/dms/upload_excel_data.html')
            
            # Check the file extension and read the file accordingly
            if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
                df = pd.read_excel(file, dtype=str)
            elif file.name.endswith('.csv'):
                df = pd.read_csv(file, dtype=str)
            else:
                messages.error(request, "Unsupported file type")
                return render(request, 'templates/dms/upload_excel_data.html')

            required_columns = [
                'company_name', 'project_name', 'location_name',
                'video_start_time', 'video_end_time', 'onedrive_url',
                'userid', 'camera_angle'
            ]

            # Check if all required columns are in the DataFrame
            missing_columns = [column for column in required_columns if column not in df.columns]
            if missing_columns:
                messages.error(request, f"Missing required columns: {', '.join(missing_columns)}")
                return render(request, 'templates/dms/upload_excel_data.html')

            # Iterate over rows in the DataFrame and create CloudURI instances
            for index, row in df.iterrows():
                try:
                    company = Company.objects.get(name=row['company_name'])
                    company_id = company.id
                except Company.DoesNotExist:
                    company_id = None

                try:
                    project = Project.objects.get(name=row['project_name'])
                    project_id = project.id
                except Project.DoesNotExist:
                    project_id = None

                NewCloudURI.objects.create(
                    userid=row.get('userid'),
                    company_id=company_id,
                    project_id=project_id,

                    company_name=row.get('company_name'),                    
                    project_name=row.get('project_name'),

                    location_name=row.get('location_name'),
                    video_start_time=row.get('video_start_time'),
                    video_end_time=row.get('video_end_time'),
                    camera_angle=row.get('camera_angle'),
                    onedrive_url=row.get('onedrive_url'),
                )

            messages.success(request, "Data has been successfully uploaded to the CloudURI model")
        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")

    return render(request, 'templates/dms/upload_excel_data.html')


# def add_cloud_uri_view(request):
#     userid = request.session.get('user_id')
#     if request.method == 'POST':
#         print("****************************",request.POST)
#         video_start_time_str = request.POST.get('video_start_time')
#         video_end_time_str = request.POST.get('video_end_time')

#         video_start_time = datetime.strptime(video_start_time_str, '%m/%d/%Y %I:%M %p')
#         video_end_time = datetime.strptime(video_end_time_str, '%m/%d/%Y %I:%M %p')

#         # Create a mutable copy of request.POST
#         mutable_post = request.POST.copy()

#         # Update the mutable copy with converted date/time values
#         mutable_post['video_start_time'] = video_start_time.strftime('%Y-%m-%d %H:%M:%S')
#         mutable_post['video_end_time'] = video_end_time.strftime('%Y-%m-%d %H:%M:%S')
#         form = CloudURIForm(mutable_post)
         
    
#         if form.is_valid():     
#             print(form.cleaned_data,"<<<<<<<<<<<<<<<<<<<<<<<<<<")    
#             form.save()
#             messages.success(request, 'Video URI Added successfully.')            
#     else:
#         form = CloudURIForm()
    
#     company_list = Company.objects.filter(userid=userid)
#     projects_list=Project.objects.filter(userid=userid)
#     context={'form': form ,'userid':userid,'projects_list':projects_list,'company_list':company_list}       
#     return render(request, 'templates/dms/add_cloud_uri.html', context=context)

def add_cloud_uri_view(request):
    userid = request.session.get('user_id')

    if request.method == 'POST':
        video_start_time_str = request.POST.get('video_start_time')
        video_end_time_str = request.POST.get('video_end_time')

        video_start_time = datetime.strptime(video_start_time_str, '%m/%d/%Y %I:%M %p')
        video_end_time = datetime.strptime(video_end_time_str, '%m/%d/%Y %I:%M %p')

        mutable_post = request.POST.copy()
        mutable_post['video_start_time'] = video_start_time.strftime('%Y-%m-%d %H:%M:%S')
        mutable_post['video_end_time'] = video_end_time.strftime('%Y-%m-%d %H:%M:%S')

        # Fetch company_id and project_id from POST data
        company_id = mutable_post.get('company_id')
        project_id = mutable_post.get('project_id')

        # Retrieve company_name and project_name based on company_id and project_id
        try:
            company_name = Company.objects.get(id=company_id).name
        except Company.DoesNotExist:
            company_name = None
        
        try:
            project_name = Project.objects.get(id=project_id).name
        except Project.DoesNotExist:
            project_name = None

        # Update mutable_post with company_name and project_name
        mutable_post['company_name'] = company_name
        mutable_post['project_name'] = project_name

        form = NewCloudURIForm(mutable_post)

        print("&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&",form)
        if form.is_valid():
            form.save()
            messages.success(request, 'Video URI Added successfully.')
        else:
            messages.error(request, 'Error adding Video URI. Please check the form.')

    else:
        form = NewCloudURIForm()

    company_list = Company.objects.filter(userid=userid)
    projects_list = Project.objects.filter(userid=userid)
    context = {'form': form, 'userid': userid, 'projects_list': projects_list, 'company_list': company_list}

    return render(request, 'templates/dms/add_cloud_uri.html', context=context)




# def filter_cloud_uri(request):
#     if request.method == 'POST':
#         print("*******************request post method data is :", request.POST)
#         userid = request.session.get('user_id')
#         if not userid:
#             return HttpResponseBadRequest("User not authenticated")

#         # company_name = request.POST.get('company')
#         # project_name = request.POST.get('project')

#         company_id = request.POST.get('company')
#         project_id = request.POST.get('project')

#         location = request.POST.get('location')
#         camera_angle = request.POST.get('cameraAngle')
#         video_start_time = request.POST.get('videoStartTime')

#         filtercolumn = {'userid': userid}

#         if company_id:
#             filtercolumn['company_id'] = company_id
#         if project_id:
#             filtercolumn['project_id'] = project_id
#         if location:
#             filtercolumn['location_name'] = location
#         if camera_angle:
#             filtercolumn['camera_angle'] = camera_angle

#         filtered_data = NewCloudURI.objects.filter(**filtercolumn)
        
#         if video_start_time == "ALL":       
#            filtered_data = NewCloudURI.objects.filter(userid=userid, location_name=location)

            
#         else:
#             filtered_data = filtered_data.annotate(
#                 video_start_date=TruncDate('video_start_time')
#             ).filter(video_start_date=video_start_time)


#         # Convert the filtered data to a list of dictionaries
#         filtered_data_json = list(filtered_data.values(
#             'company_name', 'project_name', 'location_name',
#             'camera_angle', 'video_start_time', 'video_end_time', 'onedrive_url'
#         ))

#         print("row data . . . . . . . . . .", filtered_data_json)
#         return JsonResponse(filtered_data_json, safe=False)
#     else:
#         return HttpResponseBadRequest("Invalid request")



def filter_cloud_uri(request):
    if request.method == 'POST':
        print("*******************request post method data is :", request.POST)
        
        check_filter_data = {}
        video_start_time = None
        video_end_time = None
        
        # Process POST data to extract and format filters
        for key in request.POST:
            if key != 'csrfmiddlewaretoken':
                value = request.POST.get(key)
                print(f"{key}: {value}")
                if value:
                    if key in ['video_start_time', 'video_end_time']:
                        try:
                            datetime_obj = datetime.strptime(value, '%m/%d/%Y %I:%M %p')
                            datetime_obj = datetime_obj.replace(second=0, microsecond=0)
                            if key == 'video_start_time':
                                video_start_time = datetime_obj
                                print("(***) video_start_time", datetime_obj)
                            else:
                                video_end_time = datetime_obj
                                print("(***) video_end_time", datetime_obj)
                        except ValueError:
                            print(f"Error converting {key} value: {value}")
                    else:
                        check_filter_data[key] = value

        # Filter queryset based on datetime range if both start and end times are provided
        filtered_data = NewCloudURI.objects.all()
        
        if video_start_time and video_end_time:
            filtered_data = filtered_data.filter(video_start_time__range=(video_start_time, video_end_time))
        
        # Filter queryset based on other criteria
        if check_filter_data:
            filtered_data = filtered_data.filter(**check_filter_data)

        # Serialize filtered data to JSON
        filtered_data_json = list(filtered_data.values(
            'company_name', 'project_name', 'location_name',
            'camera_angle', 'video_start_time', 'video_end_time', 'onedrive_url'
        ))

        # Return filtered data as JSON response
        return JsonResponse(filtered_data_json, safe=False)
    
    else:
        return HttpResponseBadRequest("Invalid request")



# def filter_cloud_uri(request):
#     if request.method == 'POST':
#         print("*******************request post method data is :", request.POST)
        
#         # Initialize an empty dictionary to store filter criteria
#         check_filter_data = {}
        
#         # Iterate over POST data to populate check_filter_data
#         for key in request.POST:
#             if key != 'csrfmiddlewaretoken':  # Exclude CSRF token
#                 value = request.POST.get(key)
#                 print(f"{key}: {value}")
#                 # Only add key-value pair if value is not empty
#                 if value:
#                     if key in ['video_start_time', 'video_end_time']:
#                         # Manually convert MM/DD/YYYY HH:MM AM/PM to Python datetime object
#                         try:
#                             datetime_obj = datetime.strptime(value, '%m/%d/%Y %I:%M %p')
#                             datetime_obj = datetime_obj.replace(second=0, microsecond=0)
#                             check_filter_data[key] = datetime_obj
#                         except ValueError as e:
#                             print(f"Error converting {key} value: {value}. Error: {e}")
#                     else:
#                         check_filter_data[key] = value

#         # Check if user is authenticated
#         userid = request.session.get('user_id')
#         if not userid:
#             return HttpResponseBadRequest("User not authenticated")

#         # Filter NewCloudURI queryset based on check_filter_data
#         filtered_data = NewCloudURI.objects.all()
        
#         try:
#             # Filter based on video_start_time and video_end_time range
#             if 'video_start_time' in check_filter_data:
#                 filtered_data = filtered_data.filter(
#                     video_start_time__lte=check_filter_data['video_start_time']
#                 )
#             if 'video_end_time' in check_filter_data:
#                 filtered_data = filtered_data.filter(
#                     video_end_time__gte=check_filter_data['video_end_time']
#                 )
#         except Exception as e:
#             print(f"Error filtering queryset: {e}")

#         # Create a list of dictionaries from filtered queryset
#         filtered_data_json = list(filtered_data.values(
#             'company_name', 'project_name', 'location_name',
#             'camera_angle', 'video_start_time', 'video_end_time', 'onedrive_url'
#         ))

#         # Return filtered data as JSON response
#         return JsonResponse(filtered_data_json, safe=False)
#     else:
#         return HttpResponseBadRequest("Invalid request")






data_store = {}  
def fetch_cloud_uri(request):
    if request.method == 'GET':
        company_id = request.GET.get('company_id')
        project_id = request.GET.get('project_id')
        
        userid = request.session.get('user_id')
        data_store['userid'] = userid         

        company_name = request.GET.get('company_name')
        project_name = request.GET.get('project_name')
        location = request.GET.get('location')
        camera_angle = request.GET.get('cameraangle')
        video_start_time = request.GET.get('video_start_time')
        
        response_data = {}

        if company_id:
            data_store['company_id'] = company_id
            projects = Project.objects.filter(company_id=company_id, userid=userid)
            projects_list = [{'id': project.id, 'name': project.name} for project in projects]
            return JsonResponse({'projects': projects_list})



        if project_id:   
            data_store['project_id'] = project_id         
            locations_list = NewCloudURI.objects.filter(userid=userid,
                                                         company_id=data_store['company_id'],
                                                         project_id=project_id,
                                                         ).values_list('location_name', flat=True)
            response_data['locations'] = list(locations_list)

        if location:
            data_store['location_name'] = location
            camera_angles_list = NewCloudURI.objects.filter(userid=userid, 
                                                            company_id=data_store['company_id'],
                                                             project_id=data_store['project_id'],
                                                             location_name=location,
                                                            ).values_list('camera_angle', flat=True)
            response_data['cameraangle'] = list(camera_angles_list)

        if camera_angle:
            data_store['camera_angle'] = camera_angle
            if 'location_name' in data_store and 'project_id' in data_store and 'company_id' in data_store:
                video_start_time = NewCloudURI.objects.filter(userid=userid, 
                                                              camera_angle=camera_angle,
                                                              location_name=data_store['location_name'],
                                                              project_id=data_store['project_id'],
                                                              company_id=data_store['company_id']
                                                              ).values_list('video_start_time', flat=True)
                response_data['video_start_time'] = list(video_start_time)


        return JsonResponse(response_data)
    
    else:
        return HttpResponseBadRequest("Invalid request")

















import os
import logging
from django.conf import settings
from django.http import HttpResponse, HttpResponseServerError
from django.views.decorators.csrf import csrf_exempt
from django.db import connections
from django.utils.text import get_valid_filename

logger = logging.getLogger(__name__)

@csrf_exempt
def export_database_postgresql(request):
    try:
        db_name = settings.DATABASES['default']['NAME']
        db_user = settings.DATABASES['default']['USER']
        db_host = settings.DATABASES['default']['HOST']
        db_password = settings.DATABASES['default']['PASSWORD']

        # Set PGPASSWORD environment variable if password is required
        if db_password:
            os.environ['PGPASSWORD'] = db_password

        # Connect to the database using Django's database connection
        with connections['default'].cursor() as cursor:

            # Get all table names in the database
            cursor.execute("SELECT table_name FROM information_schema.tables WHERE table_schema = 'public'")
            table_names = [row[0] for row in cursor.fetchall()]

            # Initialize an empty list to store SQL statements
            sql_lines = []

            for table_name in table_names:
                # Generate CREATE TABLE statement
                cursor.execute(f"SELECT column_name, data_type FROM information_schema.columns WHERE table_name = '{table_name}'")
                columns = cursor.fetchall()
                create_table_sql = f"CREATE TABLE {table_name} ("
                for column in columns:
                    create_table_sql += f"{column[0]} {column[1]}, "
                create_table_sql = create_table_sql[:-2] + ");"
                sql_lines.append(create_table_sql)

                # Generate INSERT INTO statements
                cursor.execute(f"SELECT * FROM {table_name}")
                table_data = cursor.fetchall()
                for row in table_data:
                    insert_sql = f"INSERT INTO {table_name} VALUES ("
                    for value in row:
                        if isinstance(value, str):
                            insert_sql += f"'{value}', "
                        else:
                            insert_sql += f"{value}, "
                    insert_sql = insert_sql[:-2] + ");"
                    sql_lines.append(insert_sql)

        # File path and name
        file_name = 'exported_data.sql'
        file_path = os.path.join(settings.MEDIA_ROOT, file_name)  # Change to your preferred directory

        # Ensure directory exists
        os.makedirs(settings.MEDIA_ROOT, exist_ok=True)

        # Write SQL statements to a file
        with open(file_path, 'w') as f:
            for line in sql_lines:
                f.write(line + "\n")

        # Prepare HTTP response with the SQL file as a downloadable attachment
        with open(file_path, 'rb') as sql_file:
            response = HttpResponse(sql_file, content_type='application/sql')
            response['Content-Disposition'] = f'attachment; filename="{get_valid_filename(file_name)}"'
            return response

    except Exception as e:
        error_message = f'Export failed: {str(e)}'
        logger.error(error_message)
        return HttpResponseServerError(error_message)




def desired_download_excel(request):
    file_path = os.path.join('static', 'assets', 'excel_format', 'cloud_data.xlsx')
    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename='cloud_data.xlsx')




def deleteall_cloud_data_table(request):
    if request.method == 'GET':
        user_id = request.session.get('user_id')        
        if user_id is not None:
            try:                
                deleted_count, _ = NewCloudURI.objects.filter(userid=user_id).delete()
               
                return JsonResponse({'deleted_count': deleted_count, 'message': 'Cloud data deleted successfully.'})
            except Exception as e:
                return JsonResponse({'error': str(e)}, status=500)
        else:
            return JsonResponse({'error': 'User ID not found in session.'}, status=400)
    else:
        return JsonResponse({'error': 'Invalid request method.'}, status=405)


def download_cloud_data_excel(request):
    userid = request.session.get('user_id')
    filtered_data = NewCloudURI.objects.filter(userid=userid).values(
        'id', 'company_name', 'project_name', 'location_name', 
        'video_start_time', 'video_end_time', 'onedrive_url', 'userid', 'camera_angle'
    )
    
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    
    # Define headers (using exact field names)
    headers = [
        'id', 'company_name', 'project_name', 'location_name', 
        'video_start_time', 'video_end_time', 'onedrive_url', 'userid', 'camera_angle'
    ]
    ws.append(headers)
    
    # Append data rows
    for item in filtered_data:
        ws.append([
            item['id'], item['company_name'], item['project_name'], 
            item['location_name'], item['video_start_time'], item['video_end_time'], 
            item['onedrive_url'], item['userid'], item['camera_angle']
        ])
    
    # Save the workbook to a BytesIO object
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Prepare response with Excel file as attachment
    response = HttpResponse(
        excel_file.read(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="cloud_data.xlsx"'
    
    return response



def view_cloud_uri(request):  
  user_id = request.session.get('user_id')

  user_account = UserAccount.objects.get(id=user_id)
  print("*******************",user_account)
  # data = CloudURI.objects.filter(userid=user_id).order_by('-id')[:10]
  data = NewCloudURI.objects.filter(userid=user_id).order_by('-id')[:10]

  company_list = Company.objects.filter(userid=user_id) 
  return render(request, 'templates/dms/view_cloud_uri.html' ,{'data' : data,'company_list':company_list,'user_account':user_account})



class AddProject(View):
    def get(self, request, *args, **kwargs):
        user_id = request.session.get('user_id')
        company_list = Company.objects.filter(userid=user_id)
        return render(request, 'templates/dms/AddProject.html', {'userid':user_id,'company_list':company_list})

    def post(self, request, *args, **kwargs):
        user_id = request.session.get('user_id')
        company_list = Company.objects.filter(userid=user_id)
        company_id = request.POST.get('company')
        project_name = request.POST.get('name')  # Assuming 'name' corresponds to project name

        try:
            company = Company.objects.get(id=company_id, userid=user_id)
        except Company.DoesNotExist:
            company = None

        if company:
            project = Project.objects.create(userid=user_id, name=project_name, company=company)
            messages.success(request, 'Project added successfully.')
        else:
            messages.error(request, 'Company does not exist or you do not have permission to add a project.')

        return render(request, 'templates/dms/AddProject.html', {'userid':user_id,'company_list':company_list})



class ListProject(View):
    def get(self, request, *args, **kwargs):
        user_id = request.session.get('user_id')
        all_projects = Project.objects.filter(userid=user_id)        
        for project in all_projects:
            project.company_name = project.company.name  # Assuming company has a 'name' field
        
        context = {'all_projects': all_projects}
        print("**********", context)
        return render(request, 'templates/dms/ListProject.html', context=context)



@login_required
def editproj_form(request):
    user_id = request.session.get('user_id')
    all_projects=Project.objects.filter(userid=user_id)
    for project in all_projects:
        project.company_name = project.company.name
    context={'all_projects':all_projects}
    return render(request,'templates/dms/ModifyProject.html', context)




class ProjectUpdate(UpdateView):
  def get(self,request,pk_id):
    user_id = request.session.get('user_id') 
    get_project=Project.objects.get(id=pk_id)
    form=ProjForm(instance=get_project)
    context={'form':form,'userid':user_id}   
    return render(request,'templates/dms/ProjectUpdate.html', context)

  def post(self, request, pk_id):      
      user_id = request.session.get('user_id') 
      existing_Project = Project.objects.get(id=pk_id)         
      update_project_name = request.POST.get('project_name')
      messages.success(request, 'Projects details Updated successfully')
      existing_Project.name = update_project_name
      existing_Project.save() 
      get_project=Project.objects.get(id=pk_id)
      form=ProjForm(instance=get_project)
      context={'form':form,'userid':user_id}   
      return render(request,'templates/dms/ProjectUpdate.html', context)






## to add new company information entered by user
def my_form(request):
  userid = request.session.get('user_id')
  if request.method == "POST":
    form = MyForm(request.POST)

    if form.is_valid():      
      form.save()
      messages.success(request, 'Company Added successfully.')  
      return render(request, 'templates/dms/AddCompany.html', {'form': form, 'userid': userid})
    else:
      return render(request, 'templates/dms/AddCompany.html', {'form': form,'userid':userid})
  else:
      form = MyForm()
  return render(request, 'templates/dms/AddCompany.html', {'form': form,'userid':userid})


## listing company details
class ListCompany(View):  
  def get(self,request,*args,**kwargs): 
    user_id = request.session.get('user_id')
    data = Company.objects.filter(userid=user_id)
    return render(request, 'templates/dms/ListCompany.html' ,{'data' : data})


def Project_delete(request, pk_id):
  print("Project delete function is working fine. . . ")
  user_id = request.session.get('user_id')

  if not user_id:
    return JsonResponse({'error' : 'You need to be logged in to delete a project'} , status=401)

  if request.method == 'GET':
    project_row_data= get_object_or_404(Project , id=pk_id)
    project_row_data.delete()
    return JsonResponse({'success': f"Project '{project_row_data.name}' has been deleted successfully "})

  all_projects=Project.objects.filter(userid=user_id)
  context={'all_projects':all_projects}
  return render(request,'templates/dms/ModifyProject.html', context)



def Company_delete(request, pk_id):
    print("Yes, It is working fine . . . ")
    user_id = request.session.get('user_id')
    if not user_id:
        return JsonResponse({'error': 'You need to be logged in to delete a company.'}, status=401)

    if request.method == 'GET':
        company = get_object_or_404(Company, id=pk_id)


        projects_to_delete = Project.objects.filter(userid=user_id, company_id=company.id)
        projects_to_delete.delete()


        company.delete()
        return JsonResponse({'success': f"Company '{company.name}' has been deleted successfully."})

    all_companies = Company.objects.filter(userid=user_id)
    context = {'all_companies': all_companies}
    return render(request, 'templates/dms/ModifyCompany.html', context)




class CompanyUpdate(UpdateView):
  def get(self,request,pk_id):
    user_id = request.session.get('user_id')   
    get_company=Company.objects.get(id=pk_id)
    form=MyForm(instance=get_company)
    context={'form':form,'userid':user_id}   
    return render(request,'templates/dms/CompanyUpdate.html', context)

  def post(self, request, pk_id):
        user_id = request.session.get('user_id')    
        update_company = get_object_or_404(Company, id=pk_id)
        form = MyForm(request.POST, instance=update_company)
        if form.is_valid():
            form.save()            
            messages.success(request, 'Company details Updated successfully')            
        else:
            print("Form is not valid")
            print(form.errors)
        return render(request, 'templates/dms/CompanyUpdate.html', {'form': form})


def dashboarddata(request):
  if request.method =="GET":
    total_company=Company.objects.all().count()
    total_projects=Project.objects.all().count()
    total_locations=Location.objects.all().count()
    total_videos=VideoFiles.objects.all().count()
    context={'total_company':total_company,'total_projects':total_projects,'total_locations':total_locations,'total_videos':total_videos}
    return render(request, "templates/dms/dashboards-analytics.html", context)
  else:
    return redirect('Error_url')

### aduit log for all the actions performed
def auditdata(form_name,action_performed,acted_by):
    acted_by='Admin' ## temporary set as admin
    auditlog = AuditLog(form_name=form_name, action_performed=action_performed,acted_by=acted_by) # create new model instance
    auditlog.save() #save to db



### loading media files images
def WorkInProgress(request):
    path = settings.MEDIA_ROOT
    print(path)
    img_list = os.listdir(path + '/images')
    print(img_list)
    context = {'images' : img_list}
    return render(request, "templates/dms/WorkInProgress.html", context)



def edit_form(request):
  
  if request.method == "POST":
    print("*************************************************************IF")

    edit_company=request.POST.GET(id)
    print('edit_company',edit_company)
    try:
      update_company=Company.objects.get(id=edit_company)
      if update_company!=None:
        print("not None")
        form = MyForm(request.POST)
        if form.is_valid():
          form.save()
          ###adding to the auditlog ###
          auditdata('Company','Update','Admin')
          return redirect('success_url')
        else:
            form =  MyForm(request.POST)
            error = "Invalid data please check"
            return render(request, 'templates/dms/ModifyCompany.html', {'form': form,'error':error})
    except Exception  as e:
          return render(request, 'templates/dms/ModifyCompany.html', {'form': form,'error':e})       
  else:
    user_id = request.session.get('user_id')    
    all_companies=Company.objects.filter(userid = user_id)
    context={'all_companies':all_companies}
    return render(request,'templates/dms/ModifyCompany.html', context)




class ListStation(View):
    def get(self,request,*args,**kwargs):
        form = LocationForm()
        all_locations=Location.objects.all()
         
        context={'all_locations':all_locations,'form':form}
       
        return render (request,'templates/dms/ListStation.html',context=context)

## to add new station information entered by user
def station_form(request):
    if request.method == "POST":
      form = LocationForm(request.POST)
      if form.is_valid():
        form.save()
        ###adding to the auditlog ###
        auditdata('Station','Add','Admin')
      return redirect('success_url')
    else:
      form = LocationForm()
      return render(request, 'templates/dms/AddStation.html', {'form': form})

# to update station information - listing for user
def editstn_form(request):
  if request.method == "POST":
    form = LocationForm(request.POST)
    if form.is_valid():
      form.save()
      ###adding to the auditlog ###
      auditdata('station','update','Admin')
      return redirect('success_url')
    else:
        form = LocationForm()
        return render(request, 'templates/dms/ModifyStation.html', {'form': form})
  else:
    all_locations=Location.objects.all()
    context={'all_locations':all_locations}

    return render(request,'templates/dms/ModifyStation.html', context)

# to update specific station information edited by user

class StationUpdate(UpdateView):
  
  def render(self,request,pk_id):
    get_station=Location.objects.get(id=pk_id)
    form=LocationForm(instance=get_station)
    context={'form':form}
    return render(request,'templates/dms/StationUpdate.html',context)

  def get(self,request,pk_id):
    get_station=Location.objects.get(id=pk_id)
    form=LocationForm(instance=get_station)
    context={'form':form}

    return render(request,'templates/dms/StationUpdate.html', context)

  def post(self,request,pk_id):
    update_station=Location.objects.get(id=pk_id)
    self.form=LocationForm(request.POST,instance=update_station)
    if self.form.is_valid():
      self.form.save()
      auditdata('station','update','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
    else:
      return self.render(request)


############## camera Model #############
## list camera models

class ListcModel(View):
  
    def get(self,request, *args,**kwargs):
        form = CModelForm()
        all_cModels=CameraModel.objects.all()
        context={'all_cModels':all_cModels,'form':form}
        
        return render(request,'templates/dms/ListCameraModel.html',context=context)

## to add new camera model information entered by user
@login_required
def Cmodel_form(request):
  if request.method == "POST":
    form = CModelForm(request.POST)
    if form.is_valid():
      form.save()
      auditdata('cameraModel','Add','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
  else:
      form = CModelForm()
      return render(request, 'templates/dms/AddCameraModel.html', {'form': form})

# to update Camera Model information modified by user
@login_required
def editcModel_form(request):
  if request.method == "POST":
    form = CModelForm(request.POST)
    if form.is_valid():
      form.save()
      auditdata('cameraModel','Update','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
    else:
        form = CModelForm()
        return render(request, 'templates/dms/ModifyCameraModel.html', {'form': form})
  else:
    all_cModels=CameraModel.objects.all()
    context={'all_cModels':all_cModels}

    return render(request,'templates/dms/ModifyCameraModel.html', context)

# to update specific camera model information edited by user

class CModelUpdate(UpdateView):
  
  def render(self,request,pk_id):
    get_cModel=CameraModel.objects.get(id=pk_id)
    form=CModelForm(instance=get_cModel)
    context={'form':form}
    return render(request,'templates/dms/CModelUpdate.html',context)

  def get(self,request,pk_id):
    get_cModel=CameraModel.objects.get(id=pk_id)
    form=CModelForm(instance=get_cModel)
    context={'form':form}

    return render(request,'templates/dms/CModelUpdate.html', context)

  def post(self,request,pk_id):
    update_cModel=CameraModel.objects.get(id=pk_id)
    self.form=CModelForm(request.POST,instance=update_cModel)
    if self.form.is_valid():
      self.form.save()
      auditdata('cameraModel','Update','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
    else:
      return self.render(request)




### Camera Position ########

class ListcPosition(View):
    def get(self,request,*args,**kwargs):
        form = CPositionForm()
        all_cpositions=CameraPosition.objects.all()
        context={'all_cpositions':all_cpositions,'form':form}
      
        return render (request,'templates/dms/ListCameraPosition.html',context=context)



## to add new camera position  information entered by user
@login_required
def CPosition_form(request):
  if request.method == "POST":
    form = CPositionForm(request.POST)
    if form.is_valid():
      form.save()
      auditdata('cameraPosition','Add','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
  else:
      form = CPositionForm()
      return render(request, 'templates/dms/AddCameraPosition.html', {'form': form})


# to update station information modified by user
@login_required
def editcPosition_form(request):
  if request.method == "POST":
    form = CPositionForm(request.POST)
    if form.is_valid():
      form.save()
      auditdata('cameraPosition','Update','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
    else:
        form = CPositionForm()
        return render(request, 'templates/dms/ModifyCameraPosition.html', {'form': form})
  else:
    all_cpositions=CameraPosition.objects.all()
    context={'all_cpositions':all_cpositions}

    return render(request,'templates/dms/ModifyCameraPosition.html', context)

# to update specific camera Position information edited by user

class CPositionUpdate(UpdateView):
  
  def render(self,request,pk_id):
    get_cpos=CameraPosition.objects.get(id=pk_id)
    form=CPositionForm(instance=get_cpos)
    context={'form':form}
    return render(request,'templates/dms/CModelUpdate.html',context)

  def get(self,request,pk_id):
    get_cpos=CameraPosition.objects.get(id=pk_id)
    form=CPositionForm(instance=get_cpos)
    context={'form':form}

    return render(request,'templates/dms/CModelUpdate.html', context)

  def post(self,request,pk_id):
    update_cPos=CameraPosition.objects.get(id=pk_id)
    self.form=CPositionForm(request.POST,instance=update_cPos)
    if self.form.is_valid():
      self.form.save()
      auditdata('cameraPosition','Update','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
    else:
      return self.render(request)


############3 USER PROFILE ##############

## to add new company information entered by user
@login_required
def add_user(request):
  if request.method == "POST":
    form = UserForm(request.POST)
    if form.is_valid():
      form.save()
      auditdata('User','Add','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
  else:
      form = UserForm()
      return render(request, 'templates/dms/AddUser.html', {'form': form})


# to update specific company information modified by user

class UserProfileUpdate(UpdateView):
  
  def render(self,request,pk_id):
    get_user=User.objects.get(id=pk_id)
    form=UserForm(instance=get_user)
    context={'form':form}
    return render(request,'templates/dms/ProfileUpdate.html',context)

  def get(self,request,pk_id):
    get_user=User.objects.get(id=pk_id)
    form=UserForm(instance=get_user)
    context={'form':form}

    return render(request,'templates/dms/ProfileUpdate.html', context)

  def post(self,request,pk_id):
    update_user=User.objects.get(id=pk_id)
    self.form=UserForm(request.POST,instance=update_user)
    if self.form.is_valid():
      self.form.save()
      auditdata('User','Update','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
    else:
      return self.render(request)


## listing company details

class ListUser(View):
    def get(self,request,*args,**kwargs):
        form = UserForm()
        all_users=User.objects.all()
        context={'all_users':all_users,'form':form}
      
        return render (request,'templates/dms/ListUser.html',context=context)

## list userprofile information for editing
@login_required
def edit_user(request):
  if request.method == "POST":
    form = UserForm(request.POST)
    if form.is_valid():
      form.save()
      auditdata('User','Update','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
    else:
        form = UserForm()
        return render(request, 'templates/dms/ModifyUser.html', {'form': form})
  else:
    all_users=User.objects.all()
    context={'all_users':all_users}

    return render(request,'templates/dms/ModifyUser.html', context)


############ ROLEs ##############

## to add new role  information entered by user
@login_required
def add_role(request):
  if request.method == "POST":
    form = RoleForm(request.POST)
    if form.is_valid():
      form.save()
      auditdata('Role','Add','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
  else:
      form = RoleForm()
      return render(request, 'templates/dms/AddRole.html', {'form': form})


# to update specific  role modified by user

class RoleUpdate(UpdateView):
  
  def render(self,request,pk_id):
    get_role=Role.objects.get(id=pk_id)
    form=RoleForm(instance=get_role)
    context={'form':form}
    return render(request,'templates/dms/RoleUpdate.html',context)

  def get(self,request,pk_id):
    get_role=Role.objects.get(id=pk_id)
    form=RoleForm(instance=get_role)
    context={'form':form}

    return render(request,'templates/dms/RoleUpdate.html', context)

  def post(self,request,pk_id):
    update_role=Role.objects.get(id=pk_id)
    self.form=RoleForm(request.POST,instance=update_role)
    if self.form.is_valid():
      self.form.save()
      auditdata('Role','Update','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
    else:
      return self.render(request)


## listing role details

class ListRole(View):
    def get(self,request,*args,**kwargs):
        form = RoleForm()
        all_roles=Role.objects.all()
        context={'all_roles':all_roles,'form':form}
      
        return render (request,'templates/dms/ListRole.html',context=context)

## list role information for editing
def edit_role(request):
  if request.method == "POST":
    form = RoleForm(request.POST)
    if form.is_valid():
      form.save()
      auditdata('Role','Update','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
    else:
        form = RoleForm()
        return render(request, 'templates/dms/ModifyRole.html', {'form': form})
  else:
    all_roles=Role.objects.all()
    context={'all_roles':all_roles}

    return render(request,'templates/dms/ModifyRole.html', context)

############ Video Files ##############

## to add new video File format  information entered by user
@login_required
def add_vff(request):
  if request.method == "POST":
    form = VideoFileForm(request.POST)
    print("Add vff")
    if form.is_valid():
      videotype=request.POST.get("vtype")
      print("videotype:",videotype)
      try:
        check_vff=VideoFF.objects.get(vtype=videotype)
        print("check_vff:",check_vff)
      except ObjectDoesNotExist:
      
          form.save()
          auditdata('VideoFileFormat','Add','Admin')  ###adding to the auditlog ###
          return redirect('success_url')
      return redirect('success_url')
  else:
      form = VideoFileForm()
      return render(request, 'templates/dms/AddVFF.html', {'form': form})


# to update specific  video File format modified by user

class VFFUpdate(UpdateView):
  
  def render(self,request,pk_id):
    get_vff=VideoFF.objects.get(id=pk_id)
    form=VideoFileForm(instance=get_vff)
    context={'form':form}
    return render(request,'templates/dms/VFFUpdate.html',context)

  def get(self,request,pk_id):
    get_vff=VideoFF.objects.get(id=pk_id)
    form=VideoFileForm(instance=get_vff)
    context={'form':form}

    return render(request,'templates/dms/VFFUpdate.html', context)

  def post(self,request,pk_id):
    update_vff=VideoFF.objects.get(id=pk_id)
    self.form=VideoFileForm(request.POST,instance=update_vff)
    if self.form.is_valid():
      self.form.save()
      auditdata('VideoFileFormat','Update','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
    else:
      return self.render(request)


## listing video File format details

class ListVFF(View):
    def get(self,request,*args,**kwargs):
        form = VideoFileForm()
        all_vff=VideoFF.objects.all()
        context={'all_vff':all_vff,'form':form}
      
        return render (request,'templates/dms/ListVFF.html',context=context)

## list video File format information for editing
@login_required
def edit_vff(request):
  if request.method == "POST":
    form = VideoFileForm(request.POST)
    if form.is_valid():
      form.save()
      auditdata('VideoFileFormat','Update','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
    else:
        form = VideoFileForm()
        return render(request, 'templates/dms/ModifyVFF.html', {'form': form})
  else:
    all_vff=VideoFF.objects.all()
    context={'all_vff':all_vff}

    return render(request,'templates/dms/ModifyVFF.html', context)

############ Light Intensity  ##############

## to add new Light Intensity category  information entered by user
@login_required
def add_li(request):
  if request.method == "POST":
    form = LIForm(request.POST)
    if form.is_valid():
      form.save()
      auditdata('LightIntensity','Add','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
  else:
      form = LIForm()
      return render(request, 'templates/dms/AddLI.html', {'form': form})


# to update specific  Light Intensity  modified by user

class LIUpdate(UpdateView):
  
  def render(self,request,pk_id):
    get_li=LItypes.objects.get(id=pk_id)
    form=LIForm(instance=get_li)
    context={'form':form}
    return render(request,'templates/dms/LIUpdate.html',context)

  def get(self,request,pk_id):
    get_li=LItypes.objects.get(id=pk_id)
    form=LIForm(instance=get_li)
    context={'form':form}

    return render(request,'templates/dms/LIUpdate.html', context)

  def post(self,request,pk_id):
    update_li=LItypes.objects.get(id=pk_id)
    self.form=LIForm(request.POST,instance=update_li)
    if self.form.is_valid():
      self.form.save()
      auditdata('LightIntensity','Update','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
    else:
      return self.render(request)


## listing Light Intensity  details

class ListLI(View):
    def get(self,request,*args,**kwargs):
        form = LIForm()
        all_li=LItypes.objects.all()
        context={'all_li':all_li,'form':form}
      
        return render (request,'templates/dms/ListLI.html',context=context)

## list Light Intensity  information for editing
@login_required
def edit_li(request):
  if request.method == "POST":
    form = LIForm(request.POST)
    if form.is_valid():
      form.save()
      auditdata('LightIntensity','Update','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
    else:
        form = LIForm()
        return render(request, 'templates/dms/ModifyLI.html', {'form': form})
  else:
    all_li=LItypes.objects.all()
    context={'all_li':all_li}

    return render(request,'templates/dms/ModifyLI.html', context)
  

### Labelling

############ Label Informations  ##############

## to add new Light Intensity category  information entered by user
@login_required
def ttImportLabel(request):
  if request.method == "POST":
    form = LIForm(request.POST)
    if form.is_valid():
      form.save()
      auditdata('ImportLabel','Add','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
  else:
      form = LIForm()
      return render(request, 'templates/dms/AddLI.html', {'form': form})
  
def postgres_connect():
# Connect to the PostgreSQL database
    conn = psycopg2.connect(
        host='127.0.0.1',
        database='vfms_db',
        user='postgres',
        password='cosaimp@2020'
    )
    # cursor = conn.cursor()
    return(conn)


@login_required
def ImportLabel(request):
    if request.method == 'POST' and request.FILES['xlsx_file']:
        # Get the uploaded file
        xlsx_file = request.FILES['xlsx_file']

        # Load the XLSX file
        wb = openpyxl.load_workbook(xlsx_file)

        # Select the active sheet
        sheet = wb.active
        conn =postgres_connect()
        cursor = conn.cursor()

        # Iterate over rows in the sheet and insert into the database
        for row in sheet.iter_rows(values_only=True):
            # Assuming your XLSX file has three columns: col1, col2, col3
            # col1, col2, col3, col4, col5, col6, col7, col8 = row
            # print("col1 : ",col1, " Col2 : ", col2 ," col3: ", col3)
            # query='INSERT INTO vfms_LabelInformation ("LName", "AnnotationCnt", "ImageCnt", "LightCdn", "CameraHeight", "RoadType", "Reviewed", "DatasetLocation") VALUES (%s,%s,%s,%s,%s,%s,%s,%s)'
            # record_to_insert=(col1, col2, col3, col4, col5, col6, col7, col8)
            col1, col2, col3, col4, col5, col6, col7, col8,col9,col10, col11, col12, col13, col14, col15, col16, col17, col18, col19, col20=row
            print("col1 : ", col1)
            fetch_query=f"""SELECT id FROM vfms_Location where name = '{col1}'"""
            print(fetch_query)
            cursor.execute(fetch_query)
            LNameID = cursor.fetchone()[0]
            print("LNameID : ",LNameID)

            # area_name auto_rickshaw   bus car others(earth mover) lcv mini bus    Two Wheeler multi axle  Truck_3Axle Truck_4Axle Truck_6Axle tracktor    tracktor_trailer    Truck_2Axle van Eicher  instance_count  image_count
            ins_query='INSERT INTO vfms_AddLabelInfo("LName_id", "Threewheelercnt", "Buscnt", "Carcnt","Otherscnt","Lcvcnt","MiniBuscnt","Twowheelercnt", "MultiAxlecnt","Axle3cnt","Axle4cnt","Axle6cnt","Tractorcnt","TTrailercnt","Truck2Acnt","Vancnt","Echiercnt","AnnotationCnt", "ImageCnt", "DatasetLocation") VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
            record_to_insert=(LNameID, col2, col3, col4, col5, col6, col7, col8,col9,col10, col11, col12, col13, col14, col15, col16, col17, col18, col19, col20)


            print(ins_query , " ---- " ,record_to_insert)
            cursor.execute(ins_query,record_to_insert)
    #          postgres_insert_query = """ INSERT INTO mobile (ID, MODEL, PRICE) VALUES (%s,%s,%s)"""
    # record_to_insert = (5, 'One Plus 6', 950)
    # cursor.execute(postgres_insert_query, record_to_insert

        # Commit the changes and close the database connection
        conn.commit()
        cursor.close()
        conn.close()

        return render(request, 'templates/dms/success.html')
    return render(request, 'templates/dms/upload_dataset.html')


## listing label  details

class ListLabel(View):
    def get(self,request,*args,**kwargs):
        form = AddLabelForm()
        all_labels=AddLabelInfo.objects.all()
        # print(all_labels)
        context={'all_labels':all_labels,'form':form}
        
      
        return render (request,'templates/dms/ListLabel.html',context=context)
    

    
## to add new location/station label information 
@login_required
def add_label(request):
  if request.method == "POST":
    form = AddLabelForm(request.POST)
    if form.is_valid():
      form.save()
      auditdata('Labelinformation','Add','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
  else:
      print("ADD LABEL INFO")
      form = AddLabelForm()
      return render(request, 'templates/dms/AddLabelInfo.html', {'form': form})
  
## list label  information for editing
@login_required
def edit_label(request):
  if request.method == "POST":
    form = AddLabelForm(request.POST)
    if form.is_valid():
      form.save()
      auditdata('AddLabelForm','Update','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
    else:
        form = AddLabelForm()
        return render(request, 'templates/dms/ModifyLabelInfo.html', {'form': form})
  else:
    all_label=AddLabelInfo.objects.all()
    context={'all_label':all_label}

    return render(request,'templates/dms/ModifyLabelInfo.html', context)
  
# to update specific  label information  modified by user

class LabelUpdate(UpdateView):
  
  def render(self,request,pk_id):
    get_label=AddLabelInfo.objects.get(id=pk_id)
    form=AddLabelForm(instance=get_label)
    context={'form':form}
    return render(request,'templates/dms/LabelUpdate.html',context)

  def get(self,request,pk_id):
    get_label=AddLabelInfo.objects.get(id=pk_id)
    form=AddLabelForm(instance=get_label)
    context={'form':form}

    return render(request,'templates/dms/LabelUpdate.html', context)

  def post(self,request,pk_id):
    update_label=AddLabelInfo.objects.get(id=pk_id)
    self.form=AddLabelForm(request.POST,instance=update_label)
    if self.form.is_valid():
      self.form.save()
      auditdata('AddLabelInfo','Update','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
    else:
      return self.render(request)

### labelling





######## Video Files upload ##############3

class GetProjects(View):

    def get(self, request):
        id1 = request.GET.get('id', None)
        projects = Project.objects.filter(company_id = id1).values('id','name')
        
        data = {'data':list(projects)}
        print(data)
        return JsonResponse(data)

class GetLocations(View):

    def get(self, request):
        id1 = request.GET.get('id', None)
        locations = Location.objects.filter(project_id = id1).values('id','name')
        
        data = {'data':list(locations)}
        return JsonResponse(data)


############videos files @@@@@@@@@
############function to filter and pass only the valid  inputs #######
def set_if_not_none(mapping, key, value):
    print(key,'__',value)
    if value is not None and not value == "":
        mapping[key] = value
        print(mapping)

       
class ListVideos(View): ## ,pk_id
    def get(self,request,*args,**kwargs):
        
        print('request.GET', request.GET)
        form = VFiles()
        # all_Videos=VideoFiles.objects.all()
        sort_params = {}
             

       ####### Get the inputs from  filter #######
        PName=request.GET.get('ProjectName', None)
        SName=request.GET.get('StationName', None)
        LICat=request.GET.get('LICate',None)
        vidstatus=request.GET.get('videostatus',None)

      ####### call function to pass only the valid filter inputs #######
        set_if_not_none(sort_params, 'ProjectName', PName)
        set_if_not_none(sort_params, 'StationName', SName)
        set_if_not_none(sort_params, 'LICate', LICat)
        set_if_not_none(sort_params, 'videostatus', vidstatus)
        print(sort_params)
        #######pass the valid inputs filters to db model to fetch records ########
        all_Videos=VideoFiles.objects.filter(**sort_params)
            
        VideoFilter=VideoFileFilter(queryset=all_Videos)
                     
        context={'all_Videos':all_Videos,'form':form,'VideoFilter':VideoFilter}
       
        return render (request,'templates/dms/ListVideos.html',context=context)

## to add new project information entered by user
@login_required
def Videos_form(request):
  print("POSTING POSTING",request.POST)
 
  # path = settings.EXTERNAL_ROOT
  # print(path)
  # Thumbnail_list = os.listdir(path + '/ThumbNailImages')
  # print(Thumbnail_list)
  # video_list = os.listdir(path + '/UploadedVideos')
  # print(video_list)
  if request.method == "POST":
    print("POST METHOD")
    form = VFiles(request.POST,request.FILES)
    if form.is_valid():
     
      print(form)
      form.save()
      auditdata('VideoFiles','Add','Admin')  ###adding to the auditlog ###

      return redirect('success_url')
    else:
      print("INVALID INVALID FORM IS INVALID")
      print(form.errors)
  else:
      print("GET METHOD")
      # user=UserAccount.objects.filter(username=user_name).get(user_company_name)
      all_comp=Company.objects.all()
      all_cameraMod=CameraModel.objects.all()
      all_vff=VideoFF.objects.all()
      all_li=LItypes.objects.all()
      form = VFiles(initial={'RejectedReason': 'NA'})
    
      context={'all_comp':all_comp,'all_cameraMod':all_cameraMod, 'all_vff':all_vff,'all_li':all_li, 'form':form}
      
      return render(request, 'templates/dms/AddVideo.html',  context)


 
      
      

# to update videofile information 
@login_required
def editvideos_form(request):
  print("inside edit videos")
  
  if request.method == "POST":
    
    form = VFiles(request.POST)

    if form.is_valid():
      form.save()
      auditdata('VideoFiles','Update','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
    else:
        form = VFiles()
        return render(request, 'templates/dms/ModifyVideo.html', {'form': form})
  else:
    
    print("inside ELSE PART")

    # all_Videos=VideoFiles.objects.all() 
    all_Videos=VideoFiles.objects.exclude(videostatus = 'Approved' )
    context={'all_Videos':all_Videos}

    return render(request,'templates/dms/ModifyVideo.html', context)  

############################################



class VideosUpdate(UpdateView):
  def render(self,request,pk_id):
    # thumbpath = Path(settings.EXTERNAL_ROOT) 
    # thumbpath.mkdir(parents=True, exist_ok=True)
    # thumbpath_str = str(thumbpath)

    # fs=FileSystemStorage(location=thumbpath_str) 
    print("inside render function",pk_id)
    if (pk_id==0):
       get_video=VideoFiles.objects.all()
    else:
      get_video=VideoFiles.objects.get(id=pk_id)
    form=VFiles(instance=get_video)
    context={'form':form}
    return render(request,'templates/dms/VideosUpdate.html',context)

  
  def get(self,request,pk_id):
    print("inside get function VideosUpdate",pk_id)
    if (pk_id==0):
       get_video=VideoFiles.objects.all()
    else:
      get_video=VideoFiles.objects.get(id=pk_id)

    form=VFiles(instance=get_video)
    
    context={'form':form}
    return render(request,'templates/dms/VideosUpdate.html',context)

  def post(self,request,pk_id):
    update_video=VideoFiles.objects.get(id=pk_id)
    update_video.videostatus="Modified"
    print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
    self.form=VFiles(request.POST,request.FILES,instance=update_video)
    print("request.FILES", request.FILES)
    print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
    if self.form.is_valid():


        ### save updated form details      
        self.form.save()
        auditdata('VideoFiles','Update','Admin')  ###adding to the auditlog ###
        return redirect('success_url')
    else:
      if (pk_id==0):
        update_video=VideoFiles.objects.all()
      else:
        update_video=VideoFiles.objects.get(id=pk_id)
      self.form=VFiles(request.POST,instance=update_video)
      print("not valid", self.form)
      return self.render(request,pk_id)
    
@login_required
def VideosApprovalRejection(request):
  print("inside VideosApprovalRejection ",request.method)
  print("inside VideosApprovalRejection ",request.FILES)
  if request.method=="POST":
      id1 = request.POST.get('id', None)
      print("RECORD ID",id1)
      vidstatus=request.POST.get("videostatus")
      print("VIDEO STATUS",vidstatus)
      if vidstatus == "Rejected":
        print("inside Rejected ")
        RejectedReason=request.POST.get("RejectedReason")
        updatevideodet1 = VideoFiles.objects.filter(id = id1).update(videostatus=vidstatus,RejectedReason=RejectedReason)
        auditdata('VideoFiles','Rejected','Admin')  ###adding to the auditlog ###
        print('updatevideodet',updatevideodet1)
      if vidstatus == "Approved":
        print("inside Approval ")
        updatevideodet2 = VideoFiles.objects.filter(id = id1).update(videostatus=vidstatus)
        auditdata('VideoFiles','Rejected','Admin')  ###adding to the auditlog ###
        print('updatevideodet',updatevideodet2)
      # updatevideodet.videostatus=request.POST.get("videostatus")
      
      
      # print('getting videostatus',request.POST.get('videostatus'))
      
      form = VFiles()
      all_Videos = VideoFiles.objects.all()
      context={'all_Videos':all_Videos,'form':form}
      data = {'context':context}
      return render(request,'templates/dms/ApproveRejectVideo.html',context)
      # return JsonResponse(data)
      # return redirect('success_url')
  else:
    print('GET GET GET')
    sort_params={}
    set_if_not_none(sort_params, 'videostatus', 'New')
    set_if_not_none(sort_params, 'videostatus', 'Modified')
    print(sort_params)
    all_Videos=VideoFiles.objects.filter(**sort_params)
    
    print(all_Videos.count())
    form = VFiles()
    context={'all_Videos':all_Videos,'form':form}   

    return render(request,'templates/dms/ApproveRejectVideo.html', context)  

###### dehradun file upload #####
@login_required
def dehradhun(request):
    if request.method == 'POST':  
        
        file = request.FILES['file'].read()
        fileName= request.POST['filename']
        existingPath = request.POST['existingPath']
        end = request.POST['end']
        nextSlice = request.POST['nextSlice']
        
        if file=="" or fileName=="" or existingPath=="" or end=="" or nextSlice=="":
            res = JsonResponse({'data':'Invalid Request'})
            return res
        else:
            if existingPath == 'null':
                # path = 'media/' + fileName
                path = settings.DRIVE_ROOT + fileName
                print('path',path)
                with open(path, 'wb+') as destination: 
                    destination.write(file)
                FileFolder = File()
                FileFolder.existingPath = fileName
                FileFolder.eof = end
                FileFolder.name = fileName
                FileFolder.save()
                if int(end):
                    res = {'data':'Saved Successfully','existingPath': fileName}
                else:
                    res = {'existingPath': fileName}
                return JsonResponse(res)

            else:
                # path = 'media/' + fileName
                path = settings.DRIVE_ROOT + fileName
                print('path',path)
                model_id = File.objects.get(existingPath=existingPath)
                if model_id.name == fileName:
                    if not model_id.eof:
                        with open(path, 'ab+') as destination: 
                            destination.write(file)
                        if int(end):
                            model_id.eof = int(end)
                            model_id.save()
                            # res = JsonResponse({'data':'Uploaded Successfully','existingPath':model_id.existingPath})
                            res = {'data':'Uploaded Successfully','existingPath':model_id.existingPath}
                        else:
                            # res = JsonResponse({'existingPath':model_id.existingPath})   
                            res =  {'existingPath':model_id.existingPath}
                        # return res
                        return JsonResponse(res)
                    else:
                        # res = JsonResponse({'data':'EOF found. Invalid request'})
                        # return res
                        res = {'data':'EOF found. Invalid request'}
                        return JsonResponse(res)
                else:
                    # res = JsonResponse({'data':'No such file exists in the existingPath'})
                    # return res
                    res = {'data':'No such file exists in the existingPath'}
                    return JsonResponse(res)
    return render(request, 'templates/upload.html')

@login_required
def CustVideos_form(request):
  print("POSTING POSTING",request.POST)
  
  if request.method == "POST":
        
        print(request.FILES)
        file = request.FILES['file'].read()
        fileName= request.POST['filename']

        existingPath = request.POST['existingPath']
        end = request.POST['end']
        nextSlice = request.POST['nextSlice']


        
        if file=="" or fileName=="" or existingPath=="" or end=="" or nextSlice=="":
            res = {'data':'Invalid Request'}
            return JsonResponse(res) 
        else:
            if existingPath == 'null':
                
                # path = 'media/' + fileName
                path = settings.DRIVE_ROOT + fileName
                print('path',path)
                with open(path, 'wb+') as destination: 
                    destination.write(file)
                print("##########################",request.POST)
                CompanyName = request.POST["CompanyName"]
                print("##########################",CompanyName)
                ProjectName = request.POST["ProjectName"]
                StationName = request.POST["StationName"]
                CameraNo = request.POST["CameraNo"]
                CModel = request.POST["CModel"]
                CHeight = request.POST["CHeight"]
                CView = request.POST["CView"]
                VideoTakenDate=request.POST["videodate"]
                VideoStartTime=request.POST["starttime"]
                VideoEndTime=request.POST["endtime"]
                LICate = request.POST["LICate"]
                Remarks = request.POST["Remarks"]
                VideoFormat = request.POST["VideoFormat"]
                

                if CompanyName and ProjectName and StationName and CameraNo and CModel and CHeight and CView and VideoTakenDate and VideoStartTime and VideoEndTime and LICate and VideoFormat :


                    CompanyName = Company.objects.get(id = CompanyName)
                    ProjectName = Project.objects.get(id = ProjectName)
                  
                    StationName = Location.objects.get(id=StationName)
                    CModel = CameraModel.objects.get(id = CModel)
                    VideoFormat = VideoFF.objects.get(id = VideoFormat)
                    LICate = LItypes.objects.get(id = LICate)
                    print("calling Try n EXCEPTION")
                    try:
                        FileFolder = CustVideoFiles.objects.get_or_create(
                              existingPath = fileName,
                                eof = end,
                                name = fileName,
                                CompanyName = CompanyName,
                                ProjectName =ProjectName,
                                StationName = StationName,
                                CameraNo = CameraNo,
                                CModel = CModel,
                                CHeight = CHeight,
                                CView = CView,
                                VideoTakenDate=VideoTakenDate,
                                VideoStartTime=VideoStartTime,
                                VideoEndTime=VideoEndTime,
                                LICate =LICate,
                                Remarks = Remarks,
                                VideoFormat = VideoFormat
                          )
                        FileFolder.save()
                    except:
                      print("INSIDE EXCEPTION")
                      res = {'data':'File already Exists','existingPath': fileName}
                      return JsonResponse(res)


                    
                    if int(end):
                        print("File upload end reached",fileName)
                        res = {'data':'Uploaded Successfully','existingPath': fileName}
                    else:
                        print("File upload no end",fileName)
                        res = {'existingPath': fileName}
                    return JsonResponse(res)
                else: 
                   return JsonResponse({'data': "invalid data ..please check"})

            else:
                path =  settings.DRIVE_ROOT + existingPath
                print("ELSE part : path" ,path)
                
                model_id = CustVideoFiles.objects.get(existingPath=existingPath)
                if model_id.name == fileName:
                    if not model_id.eof:
                        with open(path, 'ab+') as destination: 
                            destination.write(file)
                        if int(end):
                            model_id.eof = int(end)
                            model_id.save()
                            print("File upload completed",model_id.existingPath)
                            res = {'data':'Uploaded Successfully','existingPath':model_id.existingPath}
                        else:
                            print("File upload inprogress",model_id.existingPath)
                            res = {'existingPath':model_id.existingPath}    
                        return JsonResponse(res)
                    else:
                        res = {'data':'EOF found. Invalid request'}
                        return JsonResponse(res)
                else:
                    res ={'data':'No such file exists in the existingPath'}
                    return JsonResponse(res)
  return render(request, 'templates/dms/AddCustVideo.html')


class ListCustVideos(View): ## ,pk_id
    def get(self,request,*args,**kwargs):
        
        print('request.GET', request.GET)
        form = VFiles()
        # all_Videos=VideoFiles.objects.all()
        sort_params = {}
             

       ####### Get the inputs from  filter #######
        PName=request.GET.get('ProjectName', None)
        SName=request.GET.get('StationName', None)
        LICat=request.GET.get('LICate',None)
        vidstatus=request.GET.get('videostatus',None)

      ####### call function to pass only the valid filter inputs #######
        set_if_not_none(sort_params, 'ProjectName', PName)
        set_if_not_none(sort_params, 'StationName', SName)
        set_if_not_none(sort_params, 'LICate', LICat)
        set_if_not_none(sort_params, 'videostatus', vidstatus)
        print(sort_params)
        #######pass the valid inputs filters to db model to fetch records ########
        all_Videos=CustVideoFiles.objects.filter(**sort_params)
            
        VideoFilter=VideoFileFilter(queryset=all_Videos)
                     
        context={'all_Videos':all_Videos,'form':form,'VideoFilter':VideoFilter}
       
        return render (request,'templates/dms/ListCustVideo.html',context=context)
      

# to update customer videofile information 
@login_required
def editcustvideos_form(request):
  print("inside edit cust videos")
  
  if request.method == "POST":
    
    form = CustVideoFiles(request.POST)

    if form.is_valid():
      form.save()
      auditdata('CustVideoFiles','Update','Admin')  ###adding to the auditlog ###
      return redirect('success_url')
    else:
        form = CustVideoFiles()
        return render(request, 'templates/dms/ModifyCustVideo.html', {'form': form})
  else:
    
    print("inside ELSE PART")

    # all_Videos=VideoFiles.objects.all() 
    all_Videos=CustVideoFiles.objects.exclude(videostatus = 'Approved' )
    context={'all_Videos':all_Videos}

    return render(request,'templates/dms/ModifyCustVideo.html', context)  

############################################

##customer video updates

class VideosUpdateCust(UpdateView):
  def render(self,request,pk_id):
    # thumbpath = Path(settings.EXTERNAL_ROOT) 
    # thumbpath.mkdir(parents=True, exist_ok=True)
    # thumbpath_str = str(thumbpath)

    # fs=FileSystemStorage(location=thumbpath_str) 
    print("inside render function",pk_id)
    if (pk_id==0):
       get_video=CustVideoFiles.objects.all()
    else:
      get_video=CustVideoFiles.objects.get(id=pk_id)
    form=VCustFiles(instance=get_video)
    context={'form':form}
    return render(request,'templates/dms/VideosUpdateCust.html',context)

  
  def get(self,request,pk_id):
    print("inside get function VideosUpdateCust",pk_id)
    if (pk_id==0):
       get_video=CustVideoFiles.objects.all()
    else:
      get_video=CustVideoFiles.objects.get(id=pk_id)

    form=VCustFiles(instance=get_video)
    
    context={'form':form}
    return render(request,'templates/dms/VideosUpdateCust.html',context)

  def post(self,request,pk_id):
    update_video=CustVideoFiles.objects.get(id=pk_id)
    update_video.videostatus="Modified"
    print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
    self.form=VCustFiles(request.POST,request.FILES,instance=update_video)
    print("request.POST", request.POST)
    print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
    if self.form.is_valid():


        ### save updated form details      
        self.form.save()
        auditdata('VideoFiles','Update','Admin')  ###adding to the auditlog ###
        return redirect('success_url')
    else:
      if (pk_id==0):
        update_video=CustVideoFiles.objects.all()
      else:
        update_video=CustVideoFiles.objects.get(id=pk_id)
      self.form=VCustFiles(request.POST,instance=update_video)
      print("not valid", self.form)
      return self.render(request,pk_id)
    
@login_required
def CustVideosApprovalRejection(request):
  print("inside CustVideosApprovalRejection ",request.method)
  print("inside CustVideosApprovalRejection ",request.FILES)

  if request.method=="POST":
      id1 = request.POST.get('id', None)
      print("RECORD ID",id1)
      vidstatus=request.POST.get("videostatus")
      print("VIDEO STATUS",vidstatus)
      if vidstatus == "Rejected":
        print("inside Rejected ")
        RejectedReason=request.POST.get("RejectedReason")
        updatevideodet1 = CustVideoFiles.objects.filter(id = id1).update(videostatus=vidstatus,RejectedReason=RejectedReason)
        auditdata('CustVideoFiles','Rejected','Admin')  ###adding to the auditlog ###
        print('updatevideodet',updatevideodet1)
      if vidstatus == "Approved":
        print("inside Approval ")
        updatevideodet2 = CustVideoFiles.objects.filter(id = id1).update(videostatus=vidstatus)
        auditdata('CustVideoFiles','Rejected','Admin')  ###adding to the auditlog ###
        print('updatevideodet',updatevideodet2)
      # updatevideodet.videostatus=request.POST.get("videostatus")
      
      
      # print('getting videostatus',request.POST.get('videostatus'))
      
      form = VCustFiles()
      all_Videos = CustVideoFiles.objects.all()
      context={'all_Videos':all_Videos,'form':form}
      data = {'context':context}
      return render(request,'templates/dms/ApproveRejectCustVideo.html',context)
      # return JsonResponse(data)
      # return redirect('success_url')
  else:
    print('GET GET GET')
    # sort_params={}
    # set_if_not_none(sort_params, 'videostatus', 'New')
    # set_if_not_none(sort_params, 'videostatus', 'Modified')
    # print(sort_params)filter(id__in=[1, 5, 34, 567, 229])
    all_Videos=CustVideoFiles.objects.filter( videostatus__in = ['New','Modified','Rejected'])
    print(all_Videos)
    form = VCustFiles()
    context={'all_Videos':all_Videos,'form':form}   

    return render(request,'templates/dms/ApproveRejectCustVideo.html', context)  
  

  ############# testing #########
@login_required
def CustVideos_form1(request):
  print("POSTING POSTING",request.POST)
  
  if request.method == "POST":
        
        print(request.FILES)
        file = request.FILES['file'].read()
        fileName= request.POST['filename']

        existingPath = request.POST['existingPath']
        end = request.POST['end']
        nextSlice = request.POST['nextSlice']


        
        if file=="" or fileName=="" or existingPath=="" or end=="" or nextSlice=="":
            res = {'data':'Invalid Request'}
            return JsonResponse(res) 
        else:
            if existingPath == 'null':
                
                # path = 'media/' + fileName
                path = settings.DRIVE_ROOT + fileName
                print('path',path)
                with open(path, 'wb+') as destination: 
                    destination.write(file)
                print("##########################",request.POST)
                CompanyName = request.POST["CompanyName"]
                print("##########################",CompanyName)
                ProjectName = request.POST["ProjectName"]
                StationName = request.POST["StationName"]
                CameraNo = request.POST["CameraNo"]
                CModel = request.POST["CModel"]
                CHeight = request.POST["CHeight"]
                CView = request.POST["CView"]
                VideoTakenDate=request.POST["videodate"]
                VideoStartTime=request.POST["starttime"]
                VideoEndTime=request.POST["endtime"]
                LICate = request.POST["LICate"]
                Remarks = request.POST["Remarks"]
                VideoFormat = request.POST["VideoFormat"]
                

                if CompanyName and ProjectName and StationName and CameraNo and CModel and CHeight and CView and VideoTakenDate and VideoStartTime and VideoEndTime and LICate and VideoFormat :


                    CompanyName = Company.objects.get(id = CompanyName)
                    ProjectName = Project.objects.get(id = ProjectName)
                  
                    StationName = Location.objects.get(id=StationName)
                    CModel = CameraModel.objects.get(id = CModel)
                    VideoFormat = VideoFF.objects.get(id = VideoFormat)
                    LICate = LItypes.objects.get(id = LICate)
                    print("calling Try n EXCEPTION")
                    try:
                        FileFolder = CustVideoFiles.objects.get_or_create(
                              existingPath = fileName,
                                eof = end,
                                name = fileName,
                                CompanyName = CompanyName,
                                ProjectName =ProjectName,
                                StationName = StationName,
                                CameraNo = CameraNo,
                                CModel = CModel,
                                CHeight = CHeight,
                                CView = CView,
                                VideoTakenDate=VideoTakenDate,
                                VideoStartTime=VideoStartTime,
                                VideoEndTime=VideoEndTime,
                                LICate =LICate,
                                Remarks = Remarks,
                                VideoFormat = VideoFormat
                          )
                        FileFolder.save()
                    except:
                      print("INSIDE EXCEPTION")
                      res = {'data':'File already Exists','existingPath': fileName}
                      return JsonResponse(res)


                    
                    if int(end):
                        print("File upload end reached",fileName)
                        res = {'data':'Uploaded Successfully','existingPath': fileName}
                    else:
                        print("File upload no end",fileName)
                        res = {'existingPath': fileName}
                    return JsonResponse(res)
                else: 
                   return JsonResponse({'data': "invalid data ..please check"})

            else:
                path =  settings.DRIVE_ROOT + existingPath
                print("ELSE part : path" ,path)

                model_id = CustVideoFiles.objects.get(existingPath=existingPath)
                if model_id.name == fileName:
                    if not model_id.eof:
                        with open(path, 'ab+') as destination: 
                            destination.write(file)
                        if int(end):
                            model_id.eof = int(end)
                            model_id.save()
                            print("File upload completed",model_id.existingPath)
                            res = {'data':'Uploaded Successfully','existingPath':model_id.existingPath}
                        else:
                            print("File upload inprogress",model_id.existingPath)
                            res = {'existingPath':model_id.existingPath}    
                        return JsonResponse(res)
                    else:
                        res = {'data':'EOF found. Invalid request'}
                        return JsonResponse(res)
                else:
                    res ={'data':'No such file exists in the existingPath'}
                    return JsonResponse(res)
  return render(request, 'templates/dms/AddCustVideo1.html')


class AddCustVideoCustomized(UpdateView):
  def render(self,request,user_name):
    print("inside render function",user_name)
    form = VCustFiles()
    if (user_name):
      #  user_company=UserAccount.objects.filter('user_company_name').get(user_name=user_name)
       user_company=UserAccount.objects.filter(user_name=user_name).values_list("user_company_name", flat=True)
       print("user_company : ",user_company)
       get_user_company=Company.objects.filter( name__in = [user_company])
       context={'form':form,'user_company':get_user_company}
    else:
        get_user_company=Company.objects.all()
        context={'form':form,'user_company':get_user_company}
    
    context={'form':form}
    return render(request,'templates/dms/AddCustVideo.html',context)
  
  print("Inside AddCustVideoCustomized :")
  def get(self,request,user_name):
    print("inside get function AddCustVideoCustomized",user_name)
    form = VCustFiles()
    if (user_name):
      #  user_company=UserAccount.objects.filter('user_company_name').get(user_name=user_name)
       camera_model=CameraModel.objects.all()
       LICate=LItypes.objects.all()
       VideoFormat=VideoFF.objects.all()
       user_company=UserAccount.objects.filter(user_name=user_name).values_list("user_company_name", flat=True)
       print("user_company : ",user_company[0])
       usercomp=user_company[0]
       if (user_company[0] == "COSAI"):
          get_user_company=Company.objects.all()
          context={'form':form,"usercomp":usercomp,  'get_user_company':get_user_company,'camera_model':camera_model,"LICate":LICate,"VideoFormat" :VideoFormat}
       else:
          get_user_company=Company.objects.filter( name__in = [user_company])
          print("No.of company : ",len(get_user_company))
          user_compid=get_user_company[0]
          get_user_project=Project.objects.filter(company__in = [user_compid])
          print("No.of projects : ",len(get_user_project),"project name : ",get_user_project)
          project_count=len(get_user_project)
          if len(get_user_project)== 1:
            context={'form':form,"usercomp":usercomp,  'get_user_company':get_user_company, 'project_count':project_count,'get_user_project':get_user_project  ,'camera_model':camera_model,"LICate":LICate,"VideoFormat" :VideoFormat}
      #  print("company list :",get_user_company)
          context={'form':form,"usercomp":usercomp,  'get_user_company':get_user_company,'camera_model':camera_model,"LICate":LICate,"VideoFormat" :VideoFormat}
    
    return render(request,'templates/dms/AddCustVideo.html',context)

  def post(self,request,user_name):
    if request.method == "POST":
        
        print(request.FILES)
        file = request.FILES['file'].read()
        fileName= request.POST['filename']

        existingPath = request.POST['existingPath']
        end = request.POST['end']
        nextSlice = request.POST['nextSlice']


        
        if file=="" or fileName=="" or existingPath=="" or end=="" or nextSlice=="":
            res = {'data':'Invalid Request'}
            return JsonResponse(res) 
        else:
            if existingPath == 'null':
                
                # path = 'media/' + fileName
                path = settings.DRIVE_ROOT + fileName
                print('path',path)
                with open(path, 'wb+') as destination: 
                    destination.write(file)
                print("##########################",request.POST)
                CompanyName = request.POST["CompanyName"]
                print("##########################",CompanyName)
                ProjectName = request.POST["ProjectName"]
                StationName = request.POST["StationName"]
                CameraNo = request.POST["CameraNo"]
                CModel = request.POST["CModel"]
                CHeight = request.POST["CHeight"]
                CView = request.POST["CView"]
                VideoTakenDate=request.POST["videodate"]
                VideoStartTime=request.POST["starttime"]
                VideoEndTime=request.POST["endtime"]
                LICate = request.POST["LICate"]
                Remarks = request.POST["Remarks"]
                VideoFormat = request.POST["VideoFormat"]
                

                if CompanyName and ProjectName and StationName and CameraNo and CModel and CHeight and CView and VideoTakenDate and VideoStartTime and VideoEndTime and LICate and VideoFormat :


                    CompanyName = Company.objects.get(id = CompanyName)
                    ProjectName = Project.objects.get(id = ProjectName)
                  
                    StationName = Location.objects.get(id=StationName)
                    CModel = CameraModel.objects.get(id = CModel)
                    VideoFormat = VideoFF.objects.get(id = VideoFormat)
                    LICate = LItypes.objects.get(id = LICate)
                    print("calling Try n EXCEPTION")
                    try:
                        FileFolder = CustVideoFiles.objects.get_or_create(
                              existingPath = fileName,
                                eof = end,
                                name = fileName,
                                CompanyName = CompanyName,
                                ProjectName =ProjectName,
                                StationName = StationName,
                                CameraNo = CameraNo,
                                CModel = CModel,
                                CHeight = CHeight,
                                CView = CView,
                                VideoTakenDate=VideoTakenDate,
                                VideoStartTime=VideoStartTime,
                                VideoEndTime=VideoEndTime,
                                LICate =LICate,
                                Remarks = Remarks,
                                VideoFormat = VideoFormat
                          )
                        FileFolder.save()
                    except:
                      print("INSIDE EXCEPTION")
                      res = {'data':'File already Exists','existingPath': fileName}
                      return JsonResponse(res)


                    
                    if int(end):
                        print("File upload end reached",fileName)
                        res = {'data':'Uploaded Successfully','existingPath': fileName}
                    else:
                        print("File upload no end",fileName)
                        res = {'existingPath': fileName}
                    return JsonResponse(res)
                else: 
                   return JsonResponse({'data': "invalid data ..please check"})

            else:
                path =  settings.DRIVE_ROOT + existingPath
                print("ELSE part : path" ,path)
                
                model_id = CustVideoFiles.objects.get(existingPath=existingPath)
                if model_id.name == fileName:
                    if not model_id.eof:
                        with open(path, 'ab+') as destination: 
                            destination.write(file)
                        if int(end):
                            model_id.eof = int(end)
                            model_id.save()
                            print("File upload completed",model_id.existingPath)
                            res = {'data':'Uploaded Successfully','existingPath':model_id.existingPath}
                        else:
                            print("File upload inprogress",model_id.existingPath)
                            res = {'existingPath':model_id.existingPath}    
                        return JsonResponse(res)
                    else:
                        res = {'data':'EOF found. Invalid request'}
                        return JsonResponse(res)
                else:
                    res ={'data':'No such file exists in the existingPath'}
                    return JsonResponse(res)
    return render(request, 'templates/dms/AddCustVideo.html')
  
  ### to detect videos uploaded by user
  # class DetectVideos(View): ## ,pk_id

@login_required     
def detect(request):
    print("detection:", request.method)
    if request.method=='GET':
        print('request.GET', request.GET)
        sort_params={}
        form = VCustFiles()
        # all_Videos=CustVideoFiles.objects.all()
         ####### Get the inputs from  filter #######
        PName=request.GET.get('ProjectName', None)
        SName=request.GET.get('StationName', None)
        ####### call function to pass only the valid filter inputs #######
        set_if_not_none(sort_params, 'ProjectName', PName)
        set_if_not_none(sort_params, 'StationName', SName)
        #######pass the valid inputs filters to db model to fetch records ########
        all_Videos=CustVideoFiles.objects.filter(**sort_params)
            
        VideoDetect=VideoDetectFilter(queryset=all_Videos)
                     
        context={'all_Videos':all_Videos,'form':form,'VideoDetect':VideoDetect}
       
        return render (request,'templates/dms/DetectVideos.html',context=context)
    else:
       ##post method
       print("post method", request.FILES)

       ##api call to track.py with videofilename ,path and date time
##ss

class detectexternal(UpdateView):
  def render(self,request,user_name):
    print("inside render function",user_name)
    form = VCustExtFiles()

    if (user_name):
      #  user_company=UserAccount.objects.filter('user_company_name').get(user_name=user_name)
       user_company=UserAccount.objects.filter(user_name=user_name).values_list("user_company_name", flat=True)
       print("user_company : ",user_company)
       get_user_company=Company.objects.filter( name__in = [user_company])
       context={'form':form,'user_company':get_user_company}
    else:
        get_user_company=Company.objects.all()
        context={'form':form,'user_company':get_user_company}
    
    context={'form':form}
    return render(request,'templates/dms/AddVideoExternal.html',context)
  
  print("Inside detectexternal :")
  def get(self,request,user_name):
    print("inside get function  detectexternal ",user_name)
    form = VCustExtFiles()
    if (user_name):
      #  user_company=UserAccount.objects.filter('user_company_name').get(user_name=user_name)
       VideoFormat=VideoFF.objects.all()
       user_company=UserAccount.objects.filter(user_name=user_name).values_list("user_company_name", flat=True)
       print("user_company : ",user_company[0])
       usercomp=user_company[0]
       if (user_company[0] == "COSAI"):
          get_user_company=Company.objects.all()
          context={'form':form,"usercomp":usercomp,  'get_user_company':get_user_company,"VideoFormat" :VideoFormat}
       else:
          get_user_company=Company.objects.filter( name__in = [user_company])
          print("No.of company : ",len(get_user_company))
          # user_compid=get_user_company[0]
          # get_user_project=Project.objects.filter(company__in = [user_compid])
          # print("No.of projects : ",len(get_user_project),"project name : ",get_user_project)
          # project_count=len(get_user_project)
          # if len(get_user_project)== 1:
          #   context={'form':form,"usercomp":usercomp,  'get_user_company':get_user_company, 'project_count':project_count,'get_user_project':get_user_project  ,"VideoFormat" :VideoFormat}
      #  print("company list :",get_user_company)
          context={'form':form,"usercomp":usercomp,  'get_user_company':get_user_company,"VideoFormat" :VideoFormat}
    
    return render(request,'templates/dms/AddVideoExternal.html',context)

  def post(self,request,user_name):
   
    if request.method == "POST"  :
        print('external video : inside post method : ', request.POST.get('urlpath'))
        urlpath= request.POST.get('urlpath')
        CompanyName = request.POST["CompanyName"]
        ProjectName = request.POST["ProjectName"]
        StationName = request.POST["StationName"]
        VideoTakenDate=request.POST["videodate"]
        VideoStartTime=request.POST["starttime"]
        VideoEndTime=request.POST["endtime"]
        Remarks = request.POST["Remarks"]
        VideoFormat = request.POST["VideoFormat"]
        if (urlpath != "NA" or len(urlpath)>2):
            # print("url save")
            CompanyName = Company.objects.get(id = CompanyName)
            ProjectName = Project.objects.get(id = ProjectName)                      
            StationName = Location.objects.get(id=StationName)                        
            VideoFormat = VideoFF.objects.get(id = VideoFormat)

            # print("CompanyName:",CompanyName,"ProjectName :",ProjectName,"StationName : ",StationName,'VideoFormat :',VideoFormat)
            # print("VideoTakenDate:",VideoTakenDate,"VideoStartTime :",VideoStartTime,"VideoEndTime : ",VideoEndTime)

            if CompanyName and ProjectName and StationName and VideoFormat:
               
              try:
                FileFolder = CustExternalFiles.objects.create(
                                      existingPath = "NA",
                                      eof = True,
                                      name = "NA",
                                      CompanyName = CompanyName,
                                      ProjectName =ProjectName,
                                      StationName = StationName,                                    
                                      VideoTakenDate=VideoTakenDate,
                                      VideoStartTime=VideoStartTime,
                                      VideoEndTime=VideoEndTime,
                                      urlpath=urlpath,
                                      Remarks = Remarks,
                                      VideoFormat = VideoFormat
                                )
                FileFolder.save()
                res = {'data':'Uploaded Successfully'}
                return JsonResponse(res)
              except IntegrityError as e:
                    print("INSIDE EXCEPTION :",urlpath )
                    if e.__cause__:
                      print(f"PostgreSQL error message: {e.__cause__.pgcode} - {e.__cause__.pgerror}")
                    res = {'data':'creation failed'}
                    return JsonResponse(res)
            else:
               print("parent enteries check")
               return JsonResponse("Invalid Form Data. Please check")
            

        else:
            # print(request.FILES)
        
          file = request.FILES['file'].read()
          fileName= request.POST['filename']

          existingPath = request.POST['existingPath']
          end = request.POST['end']
          nextSlice = request.POST['nextSlice']

          if file=="" or fileName=="" or existingPath=="" or end=="" or nextSlice=="":
                res = {'data':'Invalid Request'}
                return JsonResponse(res) 
          else:
                if existingPath == 'null':
                    
                    # path = 'media/' + fileName
                    path = settings.DRIVE_ROOT + fileName
                    print('path',path)
                    with open(path, 'wb+') as destination: 
                        destination.write(file)
                    print("##########################",request.POST)
                    CompanyName = request.POST["CompanyName"]
                    print("##########################",CompanyName)
                    ProjectName = request.POST["ProjectName"]
                    StationName = request.POST["StationName"]
                    VideoTakenDate=request.POST["videodate"]
                    VideoStartTime=request.POST["starttime"]
                    VideoEndTime=request.POST["endtime"]
                    Remarks = request.POST["Remarks"]
                    VideoFormat = request.POST["VideoFormat"]
                    
                    if CompanyName and ProjectName and StationName and VideoTakenDate and VideoStartTime and VideoEndTime and  VideoFormat :


                        CompanyName = Company.objects.get(id = CompanyName)
                        ProjectName = Project.objects.get(id = ProjectName)                      
                        StationName = Location.objects.get(id=StationName)                        
                        VideoFormat = VideoFF.objects.get(id = VideoFormat)
                        
                        print("calling Try n EXCEPTION")
                        try:
                            FileFolder = CustExternalFiles.objects.get_or_create(
                                  existingPath = fileName,
                                    eof = end,
                                    name = fileName,
                                    CompanyName = CompanyName,
                                    ProjectName =ProjectName,
                                    StationName = StationName,
                                    
                                    VideoTakenDate=VideoTakenDate,
                                    VideoStartTime=VideoStartTime,
                                    VideoEndTime=VideoEndTime,
                                    urlpath='NA',
                                    Remarks = Remarks,
                                    VideoFormat = VideoFormat
                              )
                            FileFolder.save()
                        except:
                          print("INSIDE EXCEPTION")
                          res = {'data':'File already Exists','existingPath': fileName}
                          return JsonResponse(res)


                        
                        if int(end):
                            print("File upload end reached",fileName)
                            res = {'data':'Uploaded Successfully','existingPath': fileName}
                        else:
                            print("File upload no end",fileName)
                            res = {'existingPath': fileName}
                        return JsonResponse(res)
                    else: 
                      return JsonResponse({'data': "invalid data ..please check"})

                else:
                    path =  settings.DRIVE_ROOT + existingPath
                    print("ELSE part : path" ,path)
                    
                    model_id = CustExternalFiles.objects.get(existingPath=existingPath)
                    if model_id.name == fileName:
                        if not model_id.eof:
                            with open(path, 'ab+') as destination: 
                                destination.write(file)
                            if int(end):
                                model_id.eof = int(end)
                                model_id.save()
                                print("File upload completed",model_id.existingPath)
                                res = {'data':'Uploaded Successfully','existingPath':model_id.existingPath}
                            else:
                                print("File upload inprogress",model_id.existingPath)
                                res = {'existingPath':model_id.existingPath}    
                            return JsonResponse(res)
                        else:
                            res = {'data':'EOF found. Invalid request'}
                            return JsonResponse(res)
                    else:
                        res ={'data':'No such file exists in the existingPath'}
                        return JsonResponse(res)
    
    return render(request, 'templates/dms/AddVideoExternal.html')
@login_required
def saveextvideo(request):
   if request.method=='POST':
        print("saveextvideo - request.FILES : ",request.FILES)
        # if form.is_valid():
           
        
        # data = {'data':list(projects)}
        # print(data)
        # return JsonResponse(data)
        return render(request, 'templates/dms/AddVideoExternal.html')
   
   



@login_required
def ExternalVideoDetectionStatus(request):
  if request.method=="GET":
    print('GET GET GET')
    all_Videos=FileUploadExternal.objects.exclude(videostatus="Completed")  
    
    print(all_Videos.count())
    form = FileUploadForm()
    context={'all_Videos':all_Videos,'form':form}   

    return render(request,'templates/dms/ExternalVideoDetectStatus.html', context) 
  
  
 
class ExternalVideosUpdate(UpdateView):
  def render(self,request,pk_id):
   
    print("inside render function",pk_id)
    if (pk_id==0):
       form=FileUploadForm1()
       context={'form':form,'error':'Invalid Video ID'}
       return redirect('error_url')
    else:
      get_video=FileUploadExternal.objects.get(id=pk_id)
      form=FileUploadForm1(instance=get_video)
      context={'form':form}
      return render(request,'templates/dms/ExternalVideosUpdate.html',context)

  
  def get(self,request,pk_id):
    print("inside get function ExternalVideosUpdate",pk_id)
    if (pk_id==0):
       form=FileUploadForm1()
       context={'form':form,'error':'Invalid Video ID'}
       return render(request,'templates/dms/ExternalVideosUpdate.html',context)
    else:
        get_video=FileUploadExternal.objects.get(id=pk_id)
        form=FileUploadForm1(instance=get_video)    
        context={'form':form}
        return render(request,'templates/dms/ExternalVideosUpdate.html',context)

  def post(self,request,pk_id):
    print("HERE HERE post validity check",request.FILES)
    update_video=FileUploadExternal.objects.get(id=pk_id)
    update_video.videostatus=request.POST.get("videostatus")
    CompanyName=request.POST.get("CompName")
    ProjectName=request.POST.get("ProjName")
    FileName=update_video.files
    URLpath=update_video.urlpath
    CompanyName=CompanyName.upper().strip()
    print("update_video.videostatus :",update_video.videostatus,CompanyName,ProjectName,FileName)
    self.form=FileUploadForm1(request.POST,request.FILES,instance=update_video)
    
    print("@@@",self.form)
    if self.form.is_valid():
        if (update_video.videostatus=="Completed"):
            # if (URLpath):
            if (URLpath) and len(URLpath.strip())>0:
              print("sending email alert invoked")
              recipient_list = ['meenakshidcosai@gmail.com','costm1993@gmail.com','coscmd@gmail.com']
              content= "Detection Completed for client " + CompanyName + " project " + ProjectName + " shared via URLPath "+ URLpath
              subject="Detection Completed on "  + str(time.strftime("%Y%m%d_%H%M%S")) + " for URLPath "+ URLpath
              # sending_mail(recipient_list,content,subject) 
              print("self.form save")
              ### save updated form details      
              self.form.save()
              auditdata('ExtVideoFiles','Update','Admin')  ###adding to the auditlog ###        
              return redirect('success_url')
            else:
                print("status completed")
              ## fetch sqllite db data and insert into postgres db
                filename, _ =  os.path.splitext(str(FileName))
                # filename="L12_20231115_114420_1"
                print('calling sqllite',pk_id,filename)
                dbinsert = readsqllite(pk_id,CompanyName,filename)
                if (dbinsert): 
                  print("sending email alert invoked")
                  recipient_list = ['meenakshidcosai@gmail.com','costm1993@gmail.com','coscmd@gmail.com']
                  content= " Detection Completed for client " + CompanyName + " project " + ProjectName
                  subject="Detection Completed on "  + str(time.strftime("%Y%m%d_%H%M%S"))
                  sending_mail(recipient_list,content,subject) 
                  ###send email to client with detection reports and 5 mins video
                  # send_mail_to_client(pk_id,CompanyName,filename)

                  print("self.form save")
                  ### save updated form details      
                  self.form.save()
                  auditdata('ExtVideoFiles','Update','Admin')  ###adding to the auditlog ###
                  return redirect('success_url')
                else:
                  print(self.form.errors)
                  # self.form=FileUploadForm1(request.POST,request.FILES,instance=update_video)
                  # context={"error":self.form.errors,"form":self.form}
                  return redirect('Error_url')
                  # return render(request, 'your_template.html', {'form': form, 'errors': form.errors})
                  # return render(request,'templates/dms/ExternalVideosUpdate.html',context)
        else:
            print("self.form save")
            ### save updated form details      
            self.form.save()
            auditdata('ExtVideoFiles','Update','Admin')  ###adding to the auditlog ###
            return redirect('success_url')
   
    else:
      
      print("not valid")
      print(self.form.errors)
      return redirect('Error_url')
   
class ExternalVideosReport(UpdateView):
  def render(self,request,pk_id):
    
    # thumbpath = Path(settings.EXTERNAL_ROOT) 
    # thumbpath.mkdir(parents=True, exist_ok=True)
    # thumbpath_str = str(thumbpath)

    # fs=FileSystemStorage(location=thumbpath_str) 
    print("inside render function",pk_id)
    if (pk_id==0):
       get_video=ObjectDetection.objects.all()
    else:
      get_video=ObjectDetection.objects.filter(video_id=pk_id)
    form=ObjectDetectionReport(instance=get_video)
    context={'form':form}
    return render(request,'templates/dms/ExternalVideoDetectReport.html',context)

  
  def get(self,request,pk_id):
    print("inside get function ExternalVideosReport",pk_id)
    if (pk_id==0):
       get_video=ObjectDetection.objects.all()
    else:
      get_video=ObjectDetection.objects.filter(video_id=pk_id)

    form=ObjectDetectionReport()    
    context={'form':form,'get_video':get_video}
    print("context : ",context)
    return render(request,'templates/dms/ExternalVideoDetectReport.html',context)


    
@login_required
def ExternalVideoDetectionList(request,user_name):
  print("inside ExternalVideoDetectionList ",request.method)
  print("inside ExternalVideoDetectionList ",request.FILES)
  if request.method=="POST":
      id1 = request.POST.get('id', None)
      print("RECORD ID",id1)
      vidstatus=request.POST.get("detectionstatus")
      print("VIDEO STATUS",vidstatus)
   
      
      form = FileUploadForm1()
      all_Videos = FileUploadExternal.objects.filter()
      context={'all_Videos':all_Videos,'form':form}
      # data = {'context':context}
      print("all_videos:", all_Videos)
      return render(request,'templates/dms/ExternalVideoDetectStatusList.html',context)
      # return JsonResponse(data)
      # return redirect('success_url')
  else:
    print('GET GET GET')
    if (user_name):
      #  user_company=UserAccount.objects.filter('user_company_name').get(user_name=user_name)
      user_company=UserAccount.objects.filter(user_name=user_name).values_list("user_company_name", flat=True)
       
       
      print("user_company : ",user_company)
      if (user_company[0] == 'COSAI'):
        print(" Admin")
        all_Videos=FileUploadExternal.objects.filter(videostatus="Completed")
      else:
          print(" client")
          get_user_company=Company.objects.filter( name__in = [user_company])
          print("No.of company : ",len(get_user_company))
          user_compid=get_user_company[0]
          # get_user_project=Project.objects.filter(company__in = [user_compid])
        
          all_Videos=FileUploadExternal.objects.filter(CompanyName__in = [user_compid])
    
    print("GET all_videos:", all_Videos)
    form = FileUploadForm1()
    context={'all_Videos':all_Videos,'form':form}   

    return render(request,'templates/dms/ExternalVideoDetectStatusList.html', context)  
  


### ADDED on 16th oct by Meenakshi 
### to read the sqllite file based on the video id provided 

def readsqllite(pk,CompanyName,filename):
     
    
    postgres_conn =postgres_connect()
    postgres_cursor = postgres_conn.cursor()
    
    video_id=pk
    select_qry="SELECT count(*) FROM vfms_objectdetection where video_id_id = " + str(video_id)
   
    postgres_cursor.execute(select_qry)
    result = postgres_cursor.fetchone()[0]
    # postgres_cursor.execute(f"SELECT count(*) FROM vfms_objectdetection where video_id_id = {pk};")
    print ('fetch_query  result:',result)
    if (result == 0 or result == None):
    # Connect to the SQLite database file
        
        dbfile_path= settings.DETECTIONDB_ROOT +'/'+ CompanyName +'/'
        print("dbfile_path :",dbfile_path)
        
        db_file = dbfile_path+filename +'.db'  # Replace with the actual path to your .db file
        if os.path.exists(db_file):
            print(f'The file {db_file} exists.')
            connection = sqlite3.connect(db_file)
            ## check whether detection details exists for selected video_id
            # Create a cursor object to execute SQL queries
            cursor = connection.cursor()
            table_name = 'object_detection'
            postgres_table='vfms_objectdetection'
            # Fetch the list of tables in the database
            # cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            # tables = cursor.fetchall()

            # # Print the list of tables
            # print("Tables in the database:")
            # for table in tables:
            #     print(table[0])
            #     table_name = table[0]
            #### # Step 1:Add video id field as it is not there  as of now..
            video_id='video_id'
            # cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN {video_id} INT")
            # Step 2: Update values in the new column
            update_value = pk  # Replace with the value you want to add
            cursor.execute(f"UPDATE {table_name} SET {video_id} = ? WHERE 1", (update_value,))
            # connection.commit()
            # # Step 3:Fetch and insert data into PostgreSQL
            cursor.execute(f"SELECT * FROM {table_name};")
            data = cursor.fetchall()
            # print ("data : ", data)
            #fetch data from table and insert into postgres table
            for row in data:
                    # print("row:", row)
                    insert_query = 'INSERT INTO vfms_objectdetection("id", "date_time", "vehicle_id", "vehicle_class_id", "vehicle_name", "direction", "Cross_line", "frame_number", "image_path", "video_id_id") VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
                    # columns=("id", "date_time", "vehicle_id", "vehicle_class_id", "vehicle_name", "direction", "Cross_line", "frame_number", "image_path", "video_id_id")
                    record_to_insert = (row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9])
                    # ins_query='INSERT INTO vfms_AddLabelInfo("LName_id", "Threewheelercnt", "Buscnt", "Carcnt","Otherscnt","Lcvcnt","MiniBuscnt","Twowheelercnt", "MultiAxlecnt","Axle3cnt","Axle4cnt","Axle6cnt","Tractorcnt","TTrailercnt","Truck2Acnt","Vancnt","Echiercnt","AnnotationCnt", "ImageCnt", "DatasetLocation") VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
                    # record_to_insert=(LNameID, col2, col3, col4, col5, col6, col7, col8,col9,col10, col11, col12, col13, col14, col15, col16, col17, col18, col19, col20)
                    print("insert_query:", insert_query,record_to_insert)
                    postgres_cursor.execute(insert_query,record_to_insert)
                    postgres_conn.commit()
            print("insert completed")
        # Close the sql lite cursor and connection
            cursor.close()
            connection.close()

            postgres_cursor.execute(f"SELECT count(*) FROM vfms_objectdetection where video_id_id = {pk};")
            print ("Video_count :",postgres_cursor.fetchone()[0])
            postgres_cursor.close()
            postgres_conn.close()
            return True
    
        else:
            print(f'The file {db_file} does not exist.Try later')
            postgres_cursor.close()
            postgres_conn.close()
            return False
    else: #(fetch_query > 0):
          print("Detection data already exists")
          postgres_cursor.close()
          postgres_conn.close()
          return True
        

class ReportDownload(View):
   def get(self, request, pk_id):
        print('ReportDownload')
        # source_folder = "//10.0.0.2/Servershare/meenakshi/sqllite/"
        videofile=FileUploadExternal.objects.get(id=pk_id)
        source_folder = settings.CSVREPORT_ROOT
        compname=videofile.CompanyName
        filename=videofile.files
        file, _ = os.path.splitext(str(filename))
        report_folder = os.path.join(source_folder, str(compname))
        file_name = f"{file}.xlsx"
        
        # Specify the source file path
        source_path = os.path.join(report_folder, file_name)
        print("source_path :",source_path)
        try:
            # Check if the file exists
            if not os.path.exists(source_path):
                raise Http404(f"File '{file_name}' not found in the source folder.")

            # Open the file foNot Found:r reading
            with open(source_path, 'rb') as file:
                # Create a response with the file content
                response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="{file_name}"'

                # Return the response
                return response
        except Http404 as e:
            print(f"Error: {e}")
            # Handle 404 error
            return HttpResponse(f"File not found: {e}", status=404)
        except Exception as e:
            print(f"An error occurred: {e}")
            # Handle other errors
            return HttpResponse(f"An error occurred: {e}", status=500)
        
  

from django.shortcuts import render
@login_required
def download_file(request):
    
    source_folder = settings.DETECTIONDB_ROOT
    report_Folder=source_folder+str(pk_id) +'/'
    file_name = str(pk_id) + ".xlsx"
    
    file_path = '/path/to/your/copied/file.txt'
    file_name = 'downloaded_file.txt'

    context = {
        'file_path': file_path,
        'file_name': file_name,
    }

    return render(request, 'templates/dms/ExternalVideoDetectStatusList.html', context)

@login_required
def redirect_to_folder(request, folder_path):
 # Encode the folder path for use in a URL
    encoded_folder_path = quote(folder_path)

    # Construct the relative URL without a protocol or domain
    relative_url = f'/{encoded_folder_path}'

    # Redirect the user to the specified folder
    return HttpResponseRedirect(relative_url)
@login_required
def redirect_to_folder1(request, folder_path):
    # Assuming 'folder_path' is the path of the folder you want to redirect to
    # You may need to sanitize and validate the path to prevent security issues

    # Assuming the folder_path is a valid path, you can construct the absolute URL
    # You may need to customize this based on your actual folder structure

    # Replace '192.168.1.2' with the actual IP address or hostname of the PC
    # base_url = '\\10.0.0.2'
    file='\\10.0.0.2'
    folder_path= folder_path + '/DetectionVideos/1/'
    # Construct the full URL by joining the base URL and the quoted folder path
    absolute_url = f'{file}://{quote(folder_path)}'
    # absolute_url = f'/DETECTIONVIDEO/{folder_path}/'

    # Redirect the user to the specified folder
    return HttpResponseRedirect(absolute_url)




# relative path
def get_relative_path(full_path):
    # Remove the base path and leading slashes
    relative_path = os.path.relpath(full_path, settings.DRIVE_ROOT)
    return relative_path.replace('\\', '/')
# multifile uploads.py
from django.shortcuts import render, redirect
from .forms import FileUploadForm


# def handle_file_upload(file_content, user_name):
   
#     from google.oauth2.credentials import Credentials
#     from googleapiclient.discovery import build

#     # Load credentials (replace with your actual credential loading logic)
#     credentials = Credentials.from_authorized_user_file('path/to/credentials.json', ['https://www.googleapis.com/auth/drive'])

#     # Build the Drive API service
#     drive_service = build('drive', 'v3', credentials=credentials)

#     # Create a new file in Google Drive
#     file_metadata = {'name': user_name}
#     media = drive_service.files().create(
#         body=file_metadata,
#         media_body=drive_service.files().createMediaUpload(body=file_metadata, media_body=file_content),
#     ).execute()

#     return media.get('resumableSessionUri', None)


@csrf_exempt
@require_POST
def get_resumable_uri(request):
    try:
        # Extract filename and filesize from the request
        access_token=generate_token()
        resumable_uri=initialize_drive_upload(access_token)

        # Call Google Drive API to initialize the upload
        # ... (Use your existing code or integrate the initialize_upload function)

        # Extract the resumableURI from the Google Drive API response
        # resumable_uri = 'https://www.googleapis.com/upload/drive/v3/files?uploadType=resumable'

        return JsonResponse({'resumableURI': resumable_uri})

    except Exception as e:
        error_message = f"An unexpected error occurred: {str(e)}"
        return JsonResponse({'error': error_message}, status=500)

@csrf_exempt
@require_POST
def initialize_drive_upload(request,access_token,filename):
    print("initialize_drive_upload")
    try:
        # Extract necessary information from the request
        
        mime_type = 'video/mp4'  # Adjust as needed
        parent_folder_id = '1-gc5Tv09ponupCxQu2pOE73EKk8OxUVC'  # Adjust as needed

        # Set up headers and payload for the initialization request
        headers = {
            'Content-Type': 'application/json;charset=UTF-8',
            'Authorization': f'Bearer {access_token}',  # Replace with your access token
            'X-CSRFToken': request.META.get('HTTP_X_CSRFTOKEN', ''),
        }
        print("headers",headers)

        payload = {
            'name': filename,
            'mimeType': mime_type,
            'parents': [parent_folder_id],
        }
        print("payload",payload)

        GOOGLE_DRIVE_API_URL = 'https://www.googleapis.com/upload/drive/v3/files?uploadType=resumable'
        # Make the POST request to initialize the upload and get resumable URI
        response = requests.post(GOOGLE_DRIVE_API_URL, headers=headers, json=payload)
        print("response",response)
        if response.status_code == 200:
            # Successful initialization
            resumable_uri = response.headers.get('Location')
            return(resumable_uri)
            # return JsonResponse({'resumableURI': resumable_uri})
        else:
            return(resumable_uri)
            # Handle error response
            # return JsonResponse({'error': f'Error initializing upload: {response.status_code}'}, status=500)

    except Exception as e:
        # Handle unexpected errors
        return(str(e))
        # return JsonResponse({'error': f'Unexpected error: {str(e)}'}, status=500)

# @csrf_exempt
# @require_POST
# def upload_chunk_to_drive(request,access_token):
#     try:
#         # Extract necessary information from the request
#         resumable_uri = request.headers.get('X-Resumable-URI')
#         content_range = request.headers.get('Content-Range')
#         chunk_data = request.body

#         # Set up headers for the chunk upload
#         headers = {
#             'Content-Range': content_range,
#             'Authorization': f'Bearer {access_token}',  # Replace with your access token
#             'X-CSRFToken': request.META.get('HTTP_X_CSRFTOKEN', ''),
#         }

#         # Make the PUT request to upload the chunk to Google Drive
#         response = requests.put(resumable_uri, headers=headers, data=chunk_data)

#         if response.status_code == 200:
#             # Successful chunk upload
#             return JsonResponse({'status': 'success'})
#         else:
#             # Handle error response
#             return JsonResponse({'error': f'Error uploading chunk: {response.status_code}'}, status=500)

#     except Exception as e:
#         # Handle unexpected errors
#         return JsonResponse({'error': f'Unexpected error: {str(e)}'}, status=500)

def upload_chunk_google_drive(filename):        

    print ('upload_chunk_google_drive : ',filename)
    credentials =  get_drive_service()

    service = build('drive', 'v3', credentials=credentials)

    file_metadata = {'name': filename}
    media = MediaFileUpload(filename, resumable=True)
    request = service.files().create(body=file_metadata, media_body=media)

    response = None
    while response is None:
        status, response = request.next_chunk()
        print (status)
        if status:
            progress=status.progress() * 100
            print("progress:",progress)
            # Update progress bar in the frontend
            # update_progress_bar(filename,progress)

    # return response['id']
    vlink,dwnlink=update_file_link(response['id'])
    print( vlink,dwnlink)
    return( vlink,dwnlink)
    # if update_status :

    #   return JsonResponse({"success":"file uploaded successfully"}) 
    # else :
    #     return JsonResponse({"failed":"link update failed"}) 

    
def update_file_link(response_id):
  drive_service=get_drive_service()
  print("drive_service :",drive_service)
  # Path to the downloaded service account key file
  key_path = settings.GDRIVE_ROOT + '/vmsweb-407206-217e8e5bab3b.json'

# Create credentials from the key file
  credentials = service_account.Credentials.from_service_account_file(
    key_path, scopes=['https://www.googleapis.com/auth/drive']
  )
  print("credentials :",credentials)
# Refresh the credentials to obtain an access token
  credentials.refresh(Request())
  # print("credentials.token :",credentials.token)
# Access Token
  access_token = credentials.token

  link = drive_service.files().get(
        # fileId=new_response.get("id"),
        fileId=response_id,
        fields='webViewLink'
        ).execute()
  newfile = response_id   
  file_link = f"https://drive.google.com/uc?id={newfile}"
  # update db with the below link information
  return(link['webViewLink'],file_link)
 

def update_progress_bar(filename,progress):
    # Implement logic to update progress bar in the frontend
    # You can use Django's JsonResponse to send the progress to the frontend
    
    res = {'data':f'{filename} upload inprogress',"progress":progress}
    return JsonResponse({res}) 
# pass

class upload_files(UpdateView):
  def render(self,request,user_name):
    print("inside render function",user_name)
    form = FileUploadForm()

    if (user_name):
      #  user_company=UserAccount.objects.filter('user_company_name').get(user_name=user_name)
       user_company=UserAccount.objects.filter(user_name=user_name).values_list("user_company_name", flat=True)
       print("user_company : ",user_company[0])
       if user_company[0]=='COSAI':
          get_user_company=Company.objects.all()
          
          context={'form':form,'user_company':get_user_company}
       else: 
          get_user_company=Company.objects.filter( name__in = [user_company])
          print("get_user_company : ",get_user_company)
          context={'form':form,'user_company':get_user_company}
    else:
        error='No Information about User Company found'
        context={'form':form,'error':error}

    context={'form':form}
    # return render(request,'templates/dms/upload_files.html',context)
    return render(request,'templates/dms/upload_single_bulkfile1.html',context)
  
  # def put(self,request,user_name):
  #     print("put method called")
  #     try:
  #           # Simulated file content (replace this with actual file handling logic)
  #           file_content = request.body
  #           print("request.body:",request.body)
  #           # Assuming you have a function to handle file upload to Google Drive
  #           resumable_uri = handle_file_upload(file_content, user_name)

  #           # Return success response with resumable URI
  #           return JsonResponse({'status': 'success'})
  #     except Exception as e:
  #           print("error ", e)
  #           return JsonResponse({'status': 'error', 'message': str(e)}, status=500)



  def get(self,request,user_name):
    print("inside get function upload_files",user_name)
    if (user_name=="ww" or user_name == ""):
      print("user_name wrong : ",user_name)
      # return render(request,'templates/dms/upload_files.html')
      return render(request,'templates/dms/upload_single_bulkfile1.html')
    else: 
       
        form = FileUploadForm()
        if (user_name):
          #  user_company=UserAccount.objects.filter('user_company_name').get(user_name=user_name)
          # VideoFormat=VideoFF.objects.all()
          # Extracting file formats
          file_formats = ["mp4","mkv"]

          # Combining file formats into a single string
          VideoFormat = ",".join(file_formats)

          # Print or use the result as needed
          print(VideoFormat)
          
          user_company=UserAccount.objects.filter(user_name=user_name).values_list("user_company_name", flat=True)
          print("usercomp : ",user_company)
          if user_company:
                usercomp=user_company[0]
                print("user_company:", user_company[0],"---",usercomp)
                if (user_company[0] == "COSAI"):
                    get_user_company=Company.objects.all()
                    context={'form':form,"usercomp":usercomp,  'get_user_company':get_user_company,"VideoFormat" :VideoFormat}
                else:
                    get_user_company=Company.objects.filter( name__in = [user_company])

                    print("No.of company : ",len(get_user_company)," company : ",get_user_company)         
                    context={'form':form,"usercomp":usercomp,  'get_user_company':get_user_company,"VideoFormat" :VideoFormat}
                    # return render(request,'templates/dms/upload_files.html',context)
                    return render(request,'templates/dms/upload_single_bulkfile1.html',context)
          else:
                print("user_company is empty")
                return redirect('Error_url')
        else:
                print("user is empty")
                return redirect('Error_url')
        # return render(request,'templates/dms/upload_files.html',context)
        return render(request,'templates/dms/upload_single_bulkfile1.html',context)
 
  def post(self,request,user_name):
    print("request.method:",request.method)
    print('form : ',request.POST)

    exceptionvar=False
    if request.method == 'POST':
        # Handle other form data
       
        # data = json.loads(request.body.decode('utf-8'))
        # form_data = request.POST.get('form')
        # print("form",form_data)

        # selected_files = data.get('selectedFiles')
        # print("table",selected_files)

        company_name = request.POST.get('Company')
        project_name = request.POST.get('Project')
        Location_name = request.POST.get('Location')
        GIS = request.POST.get('gisInput')
        # print("request.POST:",request.POST)
        TrafficType = request.POST.get('TrafficType')
        Remarks = request.POST.get('Remarks')
        Ndays = request.POST.get('Ndays')
        StartDate = request.POST.get('StartDate')
        EndDate = request.POST.get('EndDate')
        urlpath=request.POST.get('urlpath')
        filesingle = request.FILES.getlist('singlefileupload')

        filebulk = request.FILES.getlist('fileupload')
        tabledata = request.POST.get('fileTableModal')
        print("tabledata:",tabledata)
        print("urlpath:",urlpath)
        print("filesingle:",filesingle,len(filesingle))   
        print("filebulk:",filebulk,len(filebulk))
           
        # ... other fields
        try:
          print("company name",company_name)
          company=Company.objects.get(name__iexact=company_name.upper().strip())
          print("company object:",company)
          if company :
           try:
              print("project",project_name,company.id)
              project=Project.objects.get(company=company.id,name__iexact=project_name.upper().strip() )
           except Project.DoesNotExist:
              exceptionvar=True
              print("project does not exist")
              NewProject=Project.objects.create(
                 company=company,
                 name=project_name,
                 date_created=datetime.datetime.now(tz=timezone.utc),
                 date_updated=datetime.datetime.now(tz=timezone.utc)
              )
              NewProject.save
              print("project created")

          else:
              exceptionvar=True
              xhr = {'error':'Error - project not created'}
              return JsonResponse(xhr)  
          project=Project.objects.get(company=company.id,name__iexact=project_name.upper().strip() )
          if project :
              try:
                print("project ID",project.id)
                location=Location.objects.get(project=project.id,name__iexact=Location_name.upper().strip())
                
                # location=Location.objects.filter(project__in = [project]).filter(name__in = [Location_name])
              except Location.DoesNotExist:
                print("location does not exist")
                exceptionvar=True
                NewLocation=Location.objects.create(
                project=project,
                name=Location_name,
                date_created=datetime.datetime.now(tz=timezone.utc),
                date_updated=datetime.datetime.now(tz=timezone.utc)
              )
                NewLocation.save 
                print("location created")
          else:
              exceptionvar=True
              xhr = {'error':'Error - Location not created'}
              return JsonResponse(xhr)  
          location=Location.objects.get(project=project.id,name__iexact=Location_name.upper().strip())
          print("location id",location.id)
          if location :
                timestr=time.strftime("%Y%m%d_%H%M%S")
                
                if (filesingle != None and len(filesingle)>0):
                   print("inside filesingle")
                   files = request.FILES.getlist('singlefileupload')
                   if (files):
                      i=0
                      for file in files:
                        i += 1
                        print("filename : ",file)
                        try:
                              

                              # Get the file extension from the original filename
                              _, file_extension = os.path.splitext(file.name)
                             
                              newfilename=f"{Location_name.upper().strip()}_{timestr}_{i}{file_extension}"
                              
                              print("newfilename:",newfilename)
                              # Get the upload path using the customized filename
                              print("companyname:",company_name)
                              new_file_instance = FileUploadExternal(
                                CompanyName=company,
                                ProjectName=project,
                                StationName=location,
                                GIS=GIS,
                                TrafficType=TrafficType,
                                Remarks=Remarks,
                                Ndays=Ndays,
                                VideoStartDate=StartDate,
                                VideoEndDate=EndDate,                                
                                files= newfilename, 
                                # gdrive_view=vlink,
                                # gdrive_dwnld=dwnlink

                                                                                                    # files=os.path.join(upload_path,newfilename),  # Use the customized filename and path
                                  # ... other fields
                              )
                              new_file_instance.save() ## Save the instance
                              file_id=new_file_instance.id
                              print('file_id :',file_id)
                              # update_progress_bar(newfilename,progress='1%')
                              access_token=generate_token()
                              print("received token :",access_token)
                              # resumable_uri=initialize_drive_upload(request,access_token,newfilename)
                              # print("resumable_uri:",resumable_uri)

                              # upload_path = get_company_upload_path(company_name.upper().strip(), newfilename)
                              # # upload_path=settings.DRIVE_ROOT+'/'+ company_name.upper().strip()
                              
                              # os.makedirs(upload_path, exist_ok=True)
                              # print("upload_path : ",upload_path)
                              video_id=new_file_instance.id
                              print("file  created:",newfilename)
                              # Now move the file to the correct location
                              # try:                              
                                
                              #   with open(os.path.normpath(os.path.join(upload_path, newfilename)), 'wb') as destination:
                              #     print("destination:",destination)
                              #     for chunk in file.chunks():
                              #         destination.write(chunk)
                              #   #  upload to gdrive()               
                              #   print("Gdrive called")

                              #   vlink,dwnlink =  gdrive_user_profile(company_name,newfilename)
                              #   # update_progress_bar(file,progress='98%')
                              #   videofile=FileUploadExternal.objects.get(id=file_id)
                              #   videofile.gdrive_view=vlink
                              #   videofile.gdrive_dwnld=dwnlink
                              #   videofile.save()
                              #   progress='99%'
                                # update_progress_bar(newfilename,progress='99.95%')
                              res = {'data':f'{newfilename} upload Inititated','filename':newfilename,"access_token":access_token,"video_id":video_id}
                              # res = {'data':f'{newfilename} uploaded successfully','filename':newfilename,'progress':progress}
                              return JsonResponse(res)
                              # except Exception as e:
                              #    exceptionvar=True  
                              #    print("error while converting to string : ", e)
                              #    xhr = {'error':'Error while converting to string'}
                              #    return JsonResponse(xhr) 
                        except Exception as e:
                            exceptionvar=True
                            print("error while converting to string : ", e)
                            xhr = {'error':'Error while converting to string'}
                            return JsonResponse(xhr) 
                          
                # elif (filebulk != None and filebulk != '' and len(filebulk)>0):
                        # instead of filebulk we use selected_files
                elif (selected_files != None and selected_files != '' and len(selected_files)>0):
                    print("inside bulk upload")
                  # files = request.FILES.getlist('fileupload')
                  # if (files):  
                    i=0
                    for file_name in selected_files:
                          i += 1
                          print("filename : ",file_name)
                          # Customize the filename as needed
                          try:
                              

                              # Get the file extension from the original filename
                              _, file_extension = os.path.splitext(file_name)
                             
                              newfilename=f"{Location_name.upper().strip()}_{timestr}_{i}{file_extension}"
                              
                              print("newfilename:",newfilename)
                              # Get the upload path using the customized filename
                              print("companyname:",company_name)
                              # upload_path = get_company_upload_path(company_name.upper().strip(), newfilename)
                              # # upload_path=settings.DRIVE_ROOT+'/'+ company_name.upper().strip()
                              
                              # os.makedirs(upload_path, exist_ok=True)
                              # print("upload_path : ",upload_path)
                              # video_id=new_file_instance.id
                              # print("file  created:",newfilename)
                              # Now move the file to the correct location
                              # try:
                               
                                
                              #   with open(os.path.normpath(os.path.join(upload_path, newfilename)), 'wb') as destination:
                              #     print("destination:",destination)
                              #     for chunk in file.chunks():
                              #         destination.write(chunk)
                              #   #  upload to gdrive()               
                              #   print("Gdrive called")
                               
                              #   # vlink,dwnlink = p12_drive(company_name,newfilename)
                              #   vlink,dwnlink = gdrive_user_profile(company_name,newfilename)
                                
                              #   ## update  DB with gdrive link to view files / download
                              #   print ("vlink :",vlink )
                              #   print("dwnlink :",dwnlink)
                              new_file_instance = FileUploadExternal(
                                CompanyName=company,
                                ProjectName=project,
                                StationName=location,
                                GIS=GIS,
                                TrafficType=TrafficType,
                                Remarks=Remarks,
                                Ndays=Ndays,
                                VideoStartDate=StartDate,
                                VideoEndDate=EndDate,                                
                                files= newfilename, 
                                # gdrive_view=vlink,
                                # gdrive_dwnld=dwnlink

                                                                                                    # files=os.path.join(upload_path,newfilename),  # Use the customized filename and path
                                  # ... other fields
                              )
                              new_file_instance.save() ## Save the instance
                              res = {'data':'Files uploaded successfully'}
                              return JsonResponse(res)
                              # print("delete locally saved file from the storage path :",os.path.normpath(os.path.join(upload_path, newfilename)))
                              #   try:
                              #     os.remove(os.path.normpath(os.path.join(upload_path, newfilename)))
                              #   except Exception as e:
                              #     exceptionvar=True
                              #     res = {'data':f'Local files not deleted ,{e}'}
                              #     return JsonResponse(res)
                              # except Exception as e:
                              #   exceptionvar=True
                              #   print("error while uploading video file : ", e)
                              #   res = {'data':'error while uploading video file'+e}
                              #   return JsonResponse(res)
                          except Exception as e:
                              exceptionvar=True
                              print("error while converting to string : ", e)
                              xhr = {'error':'Error while converting to string'}
                              return JsonResponse(xhr)
                    
                    # email alert
                    if exceptionvar == False:
                        print("email alert invoked")
                        # to_addr='meenakshidcosai@gmail.com'
                        recipient_list = ['meenakshidcosai@gmail.com','costm1993@gmail.com','coscmd@gmail.com']
                        content= "New Video files " + str(i) +" got uploaded now by client  " + company_name
                        subject="New Video files uploaded on "  + timestr +" by client  " + company_name
                        try:
                          # sending_mail(recipient_list,content,subject)
                          print("email sent")
                        except Exception as e:
                          xhr = {'error':'sending email Error'}
                          return JsonResponse(xhr)
                        res = {'data':'Files uploaded successfully'}
                        return JsonResponse(res)
                    else:
                        xhr = {'error':'Exception Error'}
                        return JsonResponse(xhr)
                else:
                    if (urlpath != None and urlpath.strip() !="" and len(urlpath.strip())>0):
                                print("inside urlpath")
                                new_file_instance = FileUploadExternal(
                                              CompanyName=company,
                                              ProjectName=project,
                                              StationName=location,
                                              GIS=GIS,
                                              TrafficType=TrafficType,
                                              Remarks=Remarks,
                                              Ndays=Ndays,
                                              VideoStartDate=StartDate,
                                              VideoEndDate=EndDate,
                                              urlpath=urlpath,
                                              
                                              # files=os.path.join(upload_path,newfilename),  # Use the customized filename and path
                                              # ... other fields
                                          )

                                          # Save the instance
                                new_file_instance.save()
                                print("email alert invoked")
                        
                                recipient_list = ['meenakshidcosai@gmail.com','costm1993@gmail.com','coscmd@gmail.com']
                                content= "New Video files link " +  urlpath + " shared now by client  " + company_name
                                subject="New Video files link  shared on "  + timestr + " by client  " + company_name
                                # sending_mail(recipient_list,content,subject)
                                print("email sent")
                                res = {'data':'Data saved successfully'}
                                return JsonResponse(res)
                    else:
                        print("error while saving urlpath: ")
                        xhr = {'error':'Error while saving urlpath ' }
                        return JsonResponse(xhr)
          else:
              xhr = {'data':'Error - location not found'}
              return JsonResponse(xhr)     
        except Exception as e:
          print("error while saving : ", e)
          xhr = {'error':'Error while uploading'}
          return JsonResponse(xhr)
 
def updateDownloadLink(request):
    if request.method == "POST":
      videoid = request.POST.get('videoid')
      downloadLink= request.POST.get('downloadLink')
      videofile=FileUploadExternal.objects.get(id=videoid)
      if (videofile != None) :
            videofile.gdrive_view=downloadLink
            videofile.gdrive_dwnld=downloadLink
            videofile.save()
            res = {'data':'File updated successfully','success':'success'}
            return JsonResponse(res)
      

def generate_token():
  print("inside generate token") 

  drive_service = get_drive_service()
  print("storage usage limit")
  get_service_Account_Storage_information(drive_service)
  
# Path to the downloaded service account key file
  key_path = settings.GDRIVE_ROOT + '/vmsweb-407206-217e8e5bab3b.json'

# Create credentials from the key file
  credentials = service_account.Credentials.from_service_account_file(
    key_path, scopes=['https://www.googleapis.com/auth/drive']
  )
  print("credentials :",credentials)
# Refresh the credentials to obtain an access token
  credentials.refresh(Request())
  # print("credentials.token :",credentials.token)
# Access Token
  access_token = credentials.token
  # print("Access Token:",access_token)
  return (access_token)
  # return JsonResponse({"data":access_token})


  def post1(self,request,user_name):
    print("request.method:",request.method)
    # delete_all_files_in_folder()
    # gdrive_profile()
    # print("delete_all_files_in_folder done")
    res = {'data':'Files Deleted successfully'}
    return JsonResponse(res)
  
  def post2(self,request,user_name):
    print("request.method:",request.method)
    print('files : ',request.FILES.getlist('fileupload'))
    exceptionvar=False
    if request.method == 'POST':
        # Handle other form data
        company_name = request.POST.get('Company')
        project_name = request.POST.get('Project')
        Location_name = request.POST.get('Location')
        GIS = request.POST.get('gisInput')
        print("GIS:",GIS)
        TrafficType = request.POST.get('TrafficType')
        Remarks = request.POST.get('Remarks')
        Ndays = request.POST.get('Ndays')
        StartDate = request.POST.get('StartDate')
        EndDate = request.POST.get('EndDate')
        urlpath=request.POST.get('urlpath')
        files = request.FILES.getlist('singlefileupload')
        
        # ... other fields
        try:
          print("company name",company_name)
          company=Company.objects.get(name__iexact=company_name.upper().strip())
          print("company object:",company)
          if company :
           try:
              print("project",project_name,company.id)
              project=Project.objects.get(company=company.id,name__iexact=project_name.upper().strip() )
           except Project.DoesNotExist:
              exceptionvar=True
              print("project does not exist")
              NewProject=Project.objects.create(
                 company=company,
                 name=project_name,
                 date_created=datetime.datetime.now(tz=timezone.utc),
                 date_updated=datetime.datetime.now(tz=timezone.utc)
              )
              NewProject.save
              print("project created")
          else:
              exceptionvar=True
              xhr = {'error':'Error - project not created'}
              return JsonResponse(xhr)  
          project=Project.objects.get(company=company.id,name__iexact=project_name.upper().strip() )
          if project :
              try:
                print("project ID",project.id)
                location=Location.objects.get(project=project.id,name__iexact=Location_name.upper().strip())
                
                # location=Location.objects.filter(project__in = [project]).filter(name__in = [Location_name])
              except Location.DoesNotExist:
                print("location does not exist")
                exceptionvar=True
                NewLocation=Location.objects.create(
                project=project,
                name=Location_name,
                date_created=datetime.datetime.now(tz=timezone.utc),
                date_updated=datetime.datetime.now(tz=timezone.utc)
              )
                NewLocation.save 
                print("location created")
          else:
              exceptionvar=True
              xhr = {'error':'Error - Location not created'}
              return JsonResponse(xhr)  
          location=Location.objects.get(project=project.id,name__iexact=Location_name.upper().strip())
          print("location id",location.id)
          if location :
                timestr=time.strftime("%Y%m%d_%H%M%S")
                files = request.FILES.getlist('singlefileupload')
                if (files):  
                    i=0
                    for file in files:
                          i += 1
                          print("filename : ",file)
                          # Customize the filename as needed
                          try:
                              

                              # Get the file extension from the original filename
                              _, file_extension = os.path.splitext(file.name)
                             
                              newfilename=f"{Location_name.upper().strip()}_{timestr}_{i}{file_extension}"
                              
                              print("newfilename:",newfilename)
                              # Get the upload path using the customized filename
                              print("companyname:",company_name)
                              upload_path = get_company_upload_path(company_name.upper().strip(), newfilename)
                              # upload_path=settings.DRIVE_ROOT+'/'+ company_name.upper().strip()
                              
                              os.makedirs(upload_path, exist_ok=True)
                              print("upload_path : ",upload_path)
                              # video_id=new_file_instance.id
                              # print("file  created:",newfilename)
                              # Now move the file to the correct location
                              try:
                               
                                
                                with open(os.path.normpath(os.path.join(upload_path, newfilename)), 'wb') as destination:
                                  print("destination:",destination)
                                  for chunk in file.chunks():
                                      destination.write(chunk)
                                #  upload to gdrive()               
                                print("Gdrive called")
                               
                                # vlink,dwnlink = p12_drive(company_name,newfilename)
                                vlink,dwnlink = gdrive_user_profile(company_name,newfilename)
                                
                                ## update  DB with gdrive link to view files / download
                                print ("vlink :",vlink )
                                print("dwnlink :",dwnlink)
                                new_file_instance = FileUploadExternal(
                                CompanyName=company,
                                ProjectName=project,
                                StationName=location,
                                GIS=GIS,
                                TrafficType=TrafficType,
                                Remarks=Remarks,
                                Ndays=Ndays,
                                VideoStartDate=StartDate,
                                VideoEndDate=EndDate,                                
                                files= newfilename, 
                                gdrive_view=vlink,
                                gdrive_dwnld=dwnlink

                                                                                                    # files=os.path.join(upload_path,newfilename),  # Use the customized filename and path
                                  # ... other fields
                              )
                                new_file_instance.save() ## Save the instance
                                print("delete locally saved file from the storage path :",os.path.normpath(os.path.join(upload_path, newfilename)))
                                try:
                                  os.remove(os.path.normpath(os.path.join(upload_path, newfilename)))
                                except Exception as e:
                                  exceptionvar=True
                                  res = {'data':f'Local files not deleted ,{e}'}
                                  return JsonResponse(res)
                              except Exception as e:
                                exceptionvar=True
                                print("error while uploading video file : ", e)
                                res = {'data':'error while uploading video file'+e}
                                return JsonResponse(res)
                          except Exception as e:
                              exceptionvar=True
                              print("error while converting to string : ", e)
                              xhr = {'error':'Error while converting to string'}
                              return JsonResponse(xhr)
                    
                    # email alert
                    if exceptionvar == False:
                        print("email alert invoked")
                        # to_addr='meenakshidcosai@gmail.com'
                        recipient_list = ['meenakshidcosai@gmail.com','costm1993@gmail.com','coscmd@gmail.com']
                        content= "New Video files " + str(i) +" got uploaded now by client  " + company_name
                        subject="New Video files uploaded on "  + timestr +" by client  " + company_name
                        try:
                          # sending_mail(recipient_list,content,subject)
                          print("email sent")
                        except Exception as e:
                          xhr = {'error':'sending email Error'}
                          return JsonResponse(xhr)
                        res = {'data':'Files uploaded successfully'}
                        return JsonResponse(res)
                    else:
                        xhr = {'error':'Exception Error'}
                        return JsonResponse(xhr)
                else:
                    if len(urlpath.strip())>0:
                                new_file_instance = FileUploadExternal(
                                              CompanyName=company,
                                              ProjectName=project,
                                              StationName=location,
                                              GIS=GIS,
                                              TrafficType=TrafficType,
                                              Remarks=Remarks,
                                              Ndays=Ndays,
                                              VideoStartDate=StartDate,
                                              VideoEndDate=EndDate,
                                              urlpath=urlpath,
                                              
                                              # files=os.path.join(upload_path,newfilename),  # Use the customized filename and path
                                              # ... other fields
                                          )

                                          # Save the instance
                                new_file_instance.save()
                                print("email alert invoked")
                        
                                recipient_list = ['meenakshidcosai@gmail.com','costm1993@gmail.com','coscmd@gmail.com']
                                content= "New Video files link " +  urlpath + " shared now by client  " + company_name
                                subject="New Video files link  shared on "  + timestr + " by client  " + company_name
                                # sending_mail(recipient_list,content,subject)
                                print("email sent")
                                res = {'data':'Data saved successfully'}
                                return JsonResponse(res)
                    else:
                        print("error while saving urlpath: ", e)
                        xhr = {'error':'Error while saving urlpath ' + e}
                        return JsonResponse(xhr)
          else:
              xhr = {'data':'Error - location not found'}
              return JsonResponse(xhr)     
        except Exception as e:
          print("error while saving : ", e)
          xhr = {'error':'Error while uploading'}
          return JsonResponse(xhr)

def send_mail_to_client(video_id,CompanyName,filename):  
# email to client with attachments:
    postgres_conn = postgres_connect()
    postgres_cursor = postgres_conn.cursor()
    
    # video_id=pk
    # Writing the query
    select_qry = """
            SELECT user_email
            FROM public.accounts_useraccount
            WHERE user_company_name = (
                SELECT name
                FROM vfms_company
                WHERE id = (
                    SELECT CompanyName_id
                    FROM vfms_fileuploadexternal
                    WHERE id = %s
                )
            )
        """

# Executing the query
    try:
      postgres_cursor.execute(select_qry,video_id)
    except Exception as e:
       print("error fetching data ", e)
       return(e)
    client_email = postgres_cursor.fetchone()[0]
    # postgres_cursor.execute(f"SELECT count(*) FROM vfms_objectdetection where video_id_id = {pk};")
    print ('fetch_query  result:',client_email)
    if (client_email == 0 or client_email == None):
       print ('no data found')
    else:
      
      to_address = client_email
      email_subject = 'Video Analytics Report'
      email_content = 'Dear Sir/Madam, \n Your video analytics report is attached for your reference \n regards \n COS AI'      
      
      csv_source_folder = settings.CSVREPORT_ROOT
      
      file, _ = os.path.splitext(str(filename))
      report_folder = csv_source_folder + '/' +str(CompanyName)+ '/' 
      report_file_name = f"{file}.xlsx"  
      detection_file_name = f"{file}.mkv"      
      # Specify the source file path
      source_path = report_folder + report_file_name
      detectionreport = source_path
      detectionvideo = settings.DETECTION_ROOT + '/' +str(CompanyName)+ '/'  + detection_file_name
      # List of file paths to be attached
      attachment_files = [detectionvideo, detectionreport]
    try:
       
      # Call the function with attachments
      sending_mail_with_attachments(to_address, email_content, email_subject, attachments=attachment_files)
    except Exception as e:
       print("error sending detection report ", e)
       return(e)
    

def get_directory_service():

  SERVICE_ACCOUNT_FILE = settings.GDRIVE_ROOT +'/vmsweb-407206-217e8e5bab3b.json'
  
  SCOPE=['https://www.googleapis.com/auth/admin.directory.user.readonly',
            'https://www.googleapis.com/auth/drive.readonly']
  credentials = service_account.Credentials.from_service_account_file(
                SERVICE_ACCOUNT_FILE, scopes=SCOPE # ['https://www.googleapis.com/auth/drive.file','https://www.googleapis.com/auth/admin.directory.user.readonly','https://www.googleapis.com/auth/admin.directory.customer.readonly','https://www.googleapis.com/auth/admin.directory.group.member.readonly']
            )
  print("credentials :",credentials)

  
  return(credentials)

def get_user_drive_storage_info(credentials,user_email):
    try:
        # Get information about the user's storage
        credentials = credentials.with_subject(user_email)
        admin_service=build('admin', 'directory_v1', credentials=credentials)
        drive_service=build('drive', 'v3', credentials=credentials)
        # user = admin_service.users().get(userKey=user_email).execute()
        user = drive_service.about().get(fields='storageQuota').execute()
        print("Storage information of the cosai domain")
               
        limit_gb = float(user['storageQuota']['limit']) / (1024 ** 3)
        usage_gb = float(user['storageQuota']['usage']) / (1024 ** 3)
        if usage_gb > 70e9:
            print('Email sent: Google Drive storage exceeds 70GB.')
            recipient_list = ['meenakshidcosai@gmail.com','costm1993@gmail.com','coscmd@gmail.com']
            content = "COSAI google drive storage exceeds 70GB. Run service to clear the storage" 
            subject = "COSAI google drive storage exceeds 70GB as of date "  + str(time.strftime("%Y%m%d_%H%M%S")) 
            sending_mail(recipient_list,content,subject) 

        print("Total Storage:", limit_gb, "GB")
        print("Used Storage:", usage_gb, "GB")
        print("Available Storage:", limit_gb - usage_gb, "GB")


    except Exception as e:
        print(f"An error occurred: {e}")

# from django.core.files.uploadedfile import TemporaryUploadedFile

# def get_file_path(temporary_uploaded_file):
#     if isinstance(temporary_uploaded_file, TemporaryUploadedFile):
#         return temporary_uploaded_file.temporary_file_path()
#     else:
#         raise ValueError("Input is not a TemporaryUploadedFile.")

def gdrive_user_profile(compname,filename):   
# Build the Admin SDK service
 ## directory service
  # directory_service = get_directory_service() #fetch 
  # print("get domain information")
  # user_email="info@cosai.in" # domain user
  # ### get storage information of the domain
  # get_user_drive_storage_info(directory_service,user_email)
 
  
  try:
    # drive_service = build('drive', 'v3', credentials=credentials)
      drive_service = get_drive_service()

  except Exception as e:
     print("error drive service",e)  
  # ### get storage information of the service account
  # get_service_Account_Storage_information(drive_service)

  ## upload video file to google drive using service account##
  folder_id='1-gc5Tv09ponupCxQu2pOE73EKk8OxUVC' ## parent folder id
  # folder_id=settings.Gdrive_Folder
  compname=compname.strip()
  filename=filename.strip()
  file_path = os.path.join(settings.DRIVE_ROOT, compname, filename)
  file_metadata = {"name":filename,"parents":[folder_id]}
  print("file_path",file_path)
  media = MediaFileUpload(file_path,chunksize=262144,resumable=True)
  request = drive_service.files().create(body=file_metadata,media_body=media,fields='id')
  print("request",request)
  response = None
  while response is None:
    status, response = request.next_chunk()
    try:
      progress=status.progress() * 100
      print("progress:",progress)
    except Exception as e:
       print("Exception : ", e)
       response = request.execute()
    # Update progress bar in the frontend
    # update_progress_bar(filename,progress)
  print("response:",response)
  response = request.execute()
  print("response1:",response)


  print(f'File uploaded successfully. File ID: {response["id"]}')
  file_id=response["id"]
  parent_id=response.get('parents',[])
  print("parent",parent_id)
  ### link to google drive files to save onto DB
  link = drive_service.files().get(
        # fileId=new_response.get("id"),
        fileId=response['id'],
        fields='webViewLink'
        ).execute()
        # print("newfile :",newfile)
  newfile=response["id"]    
  # file_link = response.get('webViewLink')
  print(link['webViewLink'])
  file_link = f"https://drive.google.com/uc?id={newfile}"
  print("file link",file_link)
  return(link['webViewLink'],file_link)

def get_service_Account_Storage_information(drive_service):
  
  try:
        # Get the service account user's storage quota information
      print("*******Service account storage details***********")
      about = drive_service.about().get(fields='storageQuota').execute()
      # Convert string values to float
      limit_gb = float(about['storageQuota']['limit']) / (1024 ** 3)
      usage_gb = float(about['storageQuota']['usage']) / (1024 ** 3)
      print("Total Storage:", limit_gb, "GB")
      print("Used Storage:", usage_gb, "GB")
      print("Available Storage:", limit_gb - usage_gb, "GB")
       # If storage exceeds 10GB, send an email alert
      if usage_gb > 10e9:
            print('Email sent: Google Drive storage exceeds 10GB.')
            recipient_list = ['meenakshidcosai@gmail.com','costm1993@gmail.com','coscmd@gmail.com']
            content = "Service account google drive storage exceeds 10GB. Run service to clear the storage" 
            subject = "Service account google drive storage exceeds 10GB as of date "  + str(time.strftime("%Y%m%d_%H%M%S")) 
            # sending_mail(recipient_list,content,subject) 
            # send_email()
            
  except Exception as e:
          print(f"Error checking storage: {e}")
          return False
# service initation
def get_drive_service():
    print("get_drive_service ")
    # GWSRoot='/datavolume/Servershare/meenakshi/DMS-latest/GWS'
    GWSRoot=''
    SERVICE_ACCOUNT_FILE = GWSRoot + '/vmsweb-407206-217e8e5bab3b.json' #settings.GDRIVE_ROOT + '/vmsweb-407206-217e8e5bab3b.json'
    print("SERVICE_ACCOUNT_FILE: ",SERVICE_ACCOUNT_FILE)
    credentials = service_account.Credentials.from_service_account_file(
                SERVICE_ACCOUNT_FILE, scopes= ['https://www.googleapis.com/auth/drive.file','https://www.googleapis.com/auth/admin.directory.user.readonly','https://www.googleapis.com/auth/admin.directory.customer.readonly','https://www.googleapis.com/auth/admin.directory.group.member.readonly']
            )
    service = build('drive', 'v3', credentials=credentials)
    return service


# Function to delete a file
def delete_file(file_id):
    drive_service = get_drive_service()
    try:
      # Delete the original file
      drive_service.files().delete(fileId=file_id).execute()
    except Exception as e:
           print(f"Error deleting original file: {e}")

def get_file_id_by_name(file_name):
    drive_service=get_drive_service()
    # Search for the file by name
    results = drive_service.files().list(q=f"name='{file_name}'").execute()
    files = results.get('files', [])

    if not files:
        print(f"No file found with the name: {file_name}")
        return None

    # Assuming you want the first file found with the given name
    return files[0]['id']           


# Function to list all files in a folder
def list_files_in_folder(service, folder_id):
    results = service.files().list(
        q=f"'{folder_id}' in parents",
        fields='files(id, name)'
    ).execute()
    files = results.get('files', [])
    return files

# Function to delete a file by ID
def delete_file(service, file_id):
    service.files().delete(fileId=file_id).execute()

# Main function to delete all files in a folder
def delete_all_files_in_folder():
    service = get_drive_service()
    folder_id='1-gc5Tv09ponupCxQu2pOE73EKk8OxUVC'
    # folder_id=settings.Gdrive_Folder
    files = list_files_in_folder(service, folder_id)

    if not files:
        print("No files found.")
    else:
        print("Deleting files:")
        for file in files:
            print(f"Deleting {file['name']} ({file['id']})")
            delete_file(service, file['id'])

# Function to check if a file is older than the specified days
def is_file_old(file, days_threshold):
    now = datetime.now()
    modified_time = datetime.strptime(file.get('modifiedTime'), "%Y-%m-%dT%H:%M:%S.%fZ")
    days_difference = (now - modified_time).days
    return days_difference > days_threshold

# Main function to delete files older than the specified days in a folder
def delete_old_files_in_folder():
    service = get_drive_service()
    folder_id='1-gc5Tv09ponupCxQu2pOE73EKk8OxUVC'
    # folder_id=settings.Gdrive_Folder
    files = list_files_in_folder(service, folder_id)

    if not files:
        print("No files found.")
    else:
        print("Deleting files older than {} days:".format(settings.DAYS_THRESHOLD))
        for file in files:
            if is_file_old(file, settings.DAYS_THRESHOLD):
                print(f"Deleting {file['name']} ({file['id']})")
                delete_file(service, file['id'])

def gdrive_profile():   
# Build the Admin SDK service
  print("\n\ndirectory service\n")
  directory_service = get_directory_service() #fetch 
  print("get domain information\n")
  user_email="info@cosai.in" # domain user
  # ### get storage information of the domain
  get_user_drive_storage_info(directory_service,user_email)
  print("\n\ndrive service\n")
  drive_service = get_drive_service()
  get_service_Account_Storage_information(drive_service)